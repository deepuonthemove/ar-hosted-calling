"""AR Voice Agent — Azure on-prem edition.

Full port of the Cloudflare Workers app (src/index.ts + src/do.ts):
  STT  Deepgram nova-2-medical  → Faster-Whisper large-v3 (local GPU)
  LLM  Workers AI / Azure OpenAI → vLLM Llama 3.1 8B (local GPU)
  TTS  Cartesia sonic-english   → Piper (local CPU)
  DB   Upstash Redis            → local Redis container
"""
import asyncio
import base64
import csv
import io
import json
import logging
import os
import re
import subprocess
import time
import datetime
import uuid
from urllib.parse import urlencode
from contextlib import asynccontextmanager

import numpy as np
import redis.asyncio as aioredis
from fastapi import FastAPI, WebSocket, Request, UploadFile, File
from fastapi.responses import HTMLResponse, Response, JSONResponse
from faster_whisper import WhisperModel
from openai import AsyncOpenAI
from openpyxl import load_workbook, Workbook

from audio import VAD, twilio_to_whisper, piper_to_twilio, rms, transcribe_live, transcribe_offline
from call_session import CallSession
from prompts import build_call_prompt, parse_markers, parse_call_result, build_greeting, load_payer, load_denial_codes, attempt_repair, strip_markers, format_dos

# ── Config ───────────────────────────────────────────────────────────────
WHISPER_MODEL_SIZE = os.getenv("WHISPER_MODEL_SIZE", "distil-large-v3")
WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cuda")
WHISPER_COMPUTE = os.getenv("WHISPER_COMPUTE", "float16")
VLLM_BASE_URL = os.getenv("VLLM_BASE_URL", "http://vllm:8001/v1")
LLM_MODEL = os.getenv("LLM_MODEL", "hugging-quants/Meta-Llama-3.1-8B-Instruct-AWQ-INT4")
LLM_MODEL_OPTIONS = {
    "Qwen 2.5 7B": "Qwen/Qwen2.5-7B-Instruct-AWQ",
    "Llama 3.1 8B": "hugging-quants/Meta-Llama-3.1-8B-Instruct-AWQ-INT4",
}
REDIS_URL = os.getenv("REDIS_URL", "redis://redis:6379")

TWILIO_ACCOUNT_SID = os.getenv("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN = os.getenv("TWILIO_AUTH_TOKEN", "")
TWILIO_FROM_NUMBER = os.getenv("TWILIO_FROM_NUMBER", "")
PUBLIC_DOMAIN = os.getenv("PUBLIC_DOMAIN", "localhost:8080")
PUBLIC_SCHEME = os.getenv("PUBLIC_SCHEME", "https")

PIPER_VOICE = os.getenv("PIPER_VOICE", "en_US-lessac-medium")
PIPER_RATE = int(os.getenv("PIPER_SAMPLE_RATE", "22050"))
PIPER_DATA_DIR = os.getenv("PIPER_DATA_DIR", "/models/piper")
# >1.0 = slower (length-scale), 1.0 = normal
PIPER_LENGTH_SCALE = float(os.getenv("PIPER_LENGTH_SCALE", "1.35"))
# Length-scale for spelled-out characters. Same lesson as Kokoro's speed:
# pushing this well above normal (was 2.0) over-stretches the vowel in each
# character's name ("es" -> a drawled "eeeese", "ay" -> "aaaaye"), hurting
# recognizability instead of helping it. Articulation is clearest at
# close-to-normal speed; the slow, clear "spelling" feel comes from the real
# silence gaps injected between characters (SPELL_CHAR_GAP_MS), not from
# stretching each word out.
PIPER_SPELL_LENGTH_SCALE = float(os.getenv("PIPER_SPELL_LENGTH_SCALE", "1.4"))

TTS_ENGINE = os.getenv("TTS_ENGINE", "piper")
KOKORO_VOICE = os.getenv("KOKORO_VOICE", "af_bella")
KOKORO_RATE = 24000
# <1.0 = slower (speed), 1.0 = normal
KOKORO_SPEED = float(os.getenv("KOKORO_SPEED", "0.81"))
KOKORO_DEVICE = os.getenv("KOKORO_DEVICE", "cuda")
# Kokoro is NOT used for spelled-out characters — measured on real audio
# that Kokoro's isolated single-word renders (e.g. "tee") come out both
# less articulate and quieter than Piper's, and slowing them down below
# ~1.0 speed breaks Kokoro's duration model outright (a bogus near-silent
# gap inside the word). Piper measured clean on the same isolated-word
# test in every dimension, so every spelled character goes through Piper
# instead, even on a call that's otherwise using Kokoro — see
# _spell_char_audio.

# Real silence injected between spelled-out characters/groups, so pacing
# doesn't depend on how each engine happens to interpret punctuation.
SPELL_CHAR_GAP_MS = int(os.getenv("SPELL_CHAR_GAP_MS", "260"))
SPELL_GROUP_GAP_MS = int(os.getenv("SPELL_GROUP_GAP_MS", "420"))

VAD_MODE = os.getenv("VAD_MODE", "silero")  # "rms" or "silero"

# Opik (LLM evaluation/tracing) — optional
OPIK_ENABLED = os.getenv("OPIK_ENABLED", "0") == "1"
OPIK_BASE_URL = os.getenv("OPIK_BASE_URL", "http://opik-backend:8080")
OPIK_API_KEY = os.getenv("OPIK_API_KEY", "")

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("ar-voice-agent")

# ── Shared state ─────────────────────────────────────────────────────────
state: dict = {}
_bg_tasks: set = set()  # strong refs for background tasks (avoids GC cancellation)


def opik_span(trace, name: str, span_type: str, input_data, output_data,
              start_ts: float | None = None, end_ts: float | None = None,
              model: str | None = None, metadata: dict | None = None):
    """Create a COMPLETE span in a single call (no separate .end()).

    Opik batches span messages; calling span.end() shortly after creation can
    drop the create payload (name/type/start_time lost → epoch-0 timestamps).
    Measuring start/end ourselves and passing them with input+output up front
    avoids that race entirely and yields a real duration.
    """
    if trace is None:
        return None
    try:
        now = datetime.datetime.now(datetime.timezone.utc)
        start = (datetime.datetime.fromtimestamp(start_ts, tz=datetime.timezone.utc)
                 if start_ts is not None else now)
        end = (datetime.datetime.fromtimestamp(end_ts, tz=datetime.timezone.utc)
               if end_ts is not None else now)
        return trace.span(name=name, type=span_type, input=input_data, output=output_data,
                          start_time=start, end_time=end,
                          model=model, provider="vllm" if span_type == "llm" else None,
                          metadata=dict(metadata or {}))
    except Exception:
        return None


def opik_start_trace(name: str, thread_id: str, input_data, metadata: dict | None = None):
    """Start a trace grouped into a thread (one per call/chat session)."""
    client = state.get("opik")
    if not client:
        return None
    try:
        return client.trace(name=name, input=input_data, metadata=metadata or {},
                            thread_id=thread_id, project_name="ar-voice-agent")
    except Exception:
        return None


def opik_end_trace(trace, output=None, error: str | None = None):
    if trace is None:
        return
    try:
        if error:
            trace.end(output={"error": error}, error_info={"error_type": "error", "message": error})
        else:
            trace.end(output=output)
    except Exception:
        pass


def load_models():
    log.info("Loading Faster-Whisper %s ...", WHISPER_MODEL_SIZE)
    state["stt_model"] = WhisperModel(WHISPER_MODEL_SIZE, device=WHISPER_DEVICE, compute_type=WHISPER_COMPUTE)
    state["stt_lock"] = asyncio.Lock()
    state["llm_client"] = AsyncOpenAI(base_url=VLLM_BASE_URL, api_key="not-needed")
    state["llm_model"] = LLM_MODEL
    state["llm_options"] = LLM_MODEL_OPTIONS
    state["redis"] = aioredis.from_url(REDIS_URL, decode_responses=True)

    # Load Piper ONNX model persistently (eliminates ~1.8s subprocess startup per utterance)
    try:
        from piper import PiperVoice
        state["piper_voice"] = PiperVoice.load(f"{PIPER_DATA_DIR}/{PIPER_VOICE}.onnx")
        log.info("PiperVoice loaded persistently (%s)", PIPER_VOICE)
    except Exception as e:
        log.warning("Persistent Piper load failed, falling back to subprocess: %s", e)
        state["piper_voice"] = None

    try:
        from kokoro import KPipeline
        state["kokoro_pipeline"] = KPipeline(lang_code='a', device=KOKORO_DEVICE)
        log.info("Kokoro TTS loaded (voice=%s, device=%s)", KOKORO_VOICE, KOKORO_DEVICE)
        # Warm up the voice + CUDA at startup so the FIRST real synthesis is
        # never delayed by a lazy voice download / GPU init.
        try:
            _ = list(state["kokoro_pipeline"]("Hello.", voice=KOKORO_VOICE, speed=KOKORO_SPEED))
        except Exception as we:
            log.warning("Kokoro warmup failed: %s", we)
    except Exception as e:
        log.warning("Kokoro not available, will fall back to Piper: %s", e)
        state["kokoro_pipeline"] = None

    # Opik LLM evaluation/tracing
    if OPIK_ENABLED:
        try:
            import opik as opik_sdk
            opik_sdk.configure(api_key=OPIK_API_KEY or "local", use_local=True,
                               url_override=OPIK_BASE_URL, project_name="ar-voice-agent")
            opik_client = opik_sdk.Opik()
            state["opik"] = opik_client
            log.info("Opik enabled at %s", OPIK_BASE_URL)
        except Exception as e:
            log.warning("Opik init failed (disabled): %s", e)
            state["opik"] = None
    else:
        state["opik"] = None

    # Preload Silero VAD (model downloads on first run)
    try:
        from audio import load_silero_vad
        model, _ = load_silero_vad()
        state["silero_loaded"] = model is not None
        log.info("Silero VAD loaded: %s", state["silero_loaded"])
    except Exception as e:
        log.warning("Silero VAD preload failed: %s", e)
        state["silero_loaded"] = False

    log.info("Models loaded.")


# ── TTS text preprocessing ───────────────────────────────────────────────
# Text transforms are shared (tts_common.py) — spelled-out runs (IDs,
# addresses, letter-by-letter names) come back wrapped in SPELL_START/
# SPELL_END sentinels. The audio layer below splits on those and synthesizes
# spelled segments character-by-character, at a slower rate, with real
# silence injected between characters — pacing that doesn't depend on how
# either engine happens to interpret punctuation.
import tts_piper as _tts_piper_mod
import tts_kokoro as _tts_kokoro_mod
from tts_common import split_spell_segments, flatten_spell_segments, SPELL_PAUSE

_TTS_MODS = {"piper": _tts_piper_mod, "kokoro": _tts_kokoro_mod}


def tts_text(text: str, engine: str = "piper") -> str:
    """Preprocess text for TTS using the engine-specific transform module."""
    return _TTS_MODS.get(engine, _tts_piper_mod).tts_text(text)


def _piper_next(gen):
    # Sentinel wrapper: StopIteration must not cross the executor/Future boundary.
    try:
        return next(gen)
    except StopIteration:
        return None


def _silence_bytes(ms: int) -> bytes:
    n = int(PIPER_RATE * ms / 1000)
    return np.zeros(n, dtype=np.int16).tobytes()


# Feeding the engine a bare, context-free symbol ("T", "5") triggers its
# "spell this out" handling, which both engines render with a leading
# percussive "t"/click — happens for digits too, and trimming to remove it
# sometimes ate the real letter instead (short letters vs. a fixed-ms trim
# can't be told apart reliably). That means the artifact is baked into the
# generated content, not a splice/boundary effect — no amount of waveform
# post-processing fixes that. So instead we don't synthesize the bare
# symbol at all: each character is spoken as its ordinary name ("tee",
# "five") — a normal short word, which is squarely inside what these
# engines are trained on and renders cleanly.
_LETTER_NAMES = {
    "A": "ay", "B": "bee", "C": "see", "D": "dee", "E": "ee", "F": "eff",
    "G": "jee", "H": "aitch", "I": "eye", "J": "jay", "K": "kay", "L": "el",
    "M": "em", "N": "en", "O": "oh", "P": "pee", "Q": "cue", "R": "are",
    "S": "es", "T": "tee", "U": "you", "V": "vee", "W": "double you",
    "X": "ex", "Y": "why", "Z": "zee",
}
_DIGIT_NAMES = {
    "0": "zero", "1": "one", "2": "two", "3": "three", "4": "four",
    "5": "five", "6": "six", "7": "seven", "8": "eight", "9": "nine",
}


def _spoken_char(tok: str) -> str:
    return _LETTER_NAMES.get(tok.upper(), _DIGIT_NAMES.get(tok, tok))


# Light safety-net fade for a clean splice into the surrounding silence.
_SPELL_FADE_MS = int(os.getenv("SPELL_FADE_MS", "12"))

# Measured on real audio: Kokoro renders very short isolated words (e.g.
# "tee" at -15dBFS, "es" at -13dBFS) noticeably quieter than normal
# continuous speech (-6 to -8dBFS) — some spelled letters end up too quiet
# to hear clearly. Boost each character's peak up to this target so it
# matches normal speech loudness. Only ever boosts UP, never attenuates —
# Piper's per-word output is already near full-scale, so this is a no-op
# for it and only helps Kokoro's quieter renders.
_SPELL_TARGET_PEAK = int(os.getenv("SPELL_TARGET_PEAK", "22000"))


def _fade_edges(pcm_bytes: bytes, fade_ms: int = _SPELL_FADE_MS,
                 target_peak: int = _SPELL_TARGET_PEAK) -> bytes:
    if not pcm_bytes:
        return pcm_bytes
    arr = np.frombuffer(pcm_bytes, dtype=np.int16).astype(np.float32)
    peak = np.abs(arr).max()
    if 0 < peak < target_peak:
        arr = np.clip(arr * (target_peak / peak), -32767, 32767)
    n = len(arr)
    fade_n = min(int(PIPER_RATE * fade_ms / 1000), n // 2)
    if fade_n > 0:
        ramp = np.linspace(0.0, 1.0, fade_n, dtype=np.float32)
        arr[:fade_n] *= ramp
        arr[-fade_n:] *= ramp[::-1]
    return arr.astype(np.int16).tobytes()


# Each spelled character is its own fresh synth call (phonemizer + model
# invocation from scratch), which can occasionally take much longer than a
# continuation chunk of normal speech would. asyncio can only deliver a
# cancellation (barge-in) at the next await checkpoint — if that checkpoint
# is buried inside one slow character call, cancellation is deferred until
# it finishes. On a long spelled ID that delay stacks up across characters
# and barge-in can feel like it isn't working. Bound each character to a
# timeout so a slow/stuck call can never block cancellation for long.
SPELL_CHAR_TIMEOUT_S = float(os.getenv("SPELL_CHAR_TIMEOUT_S", "2.5"))


async def _synth_char_bytes(char_gen, label: str = "?", engine: str = "?",
                             timeout: float = SPELL_CHAR_TIMEOUT_S) -> bytes:
    """Buffer one spelled character's audio (always short) and fade its
    edges, instead of streaming it raw — removes the click/static pop that a
    fresh isolated synth call produces at onset. Bounded by `timeout` so a
    single character can't stall cancellation indefinitely. Logs how long
    the call took, so it's visible whether per-character synth is the thing
    delaying barge-in during a long spell."""
    async def _collect():
        buf = bytearray()
        async for b in char_gen:
            buf.extend(b)
        return bytes(buf)

    t0 = time.time()
    try:
        buf = await asyncio.wait_for(_collect(), timeout=timeout)
    except asyncio.TimeoutError:
        log.warning("[spell] %s char %r: exceeded %.1fs — skipping it", engine, label, timeout)
        return b""
    elapsed_ms = (time.time() - t0) * 1000
    log.info("[spell] %s char %r: %.0fms", engine, label, elapsed_ms)
    return _fade_edges(buf)


# Serializes in-process Piper synthesis. The shared ONNX session is not safe
# under concurrency (e.g. a cancelled TTS leaving an executor thread mid-inference).
_tts_lock = None


async def _get_tts_lock():
    global _tts_lock
    if _tts_lock is None:
        _tts_lock = asyncio.Lock()
    return _tts_lock


# Measured on real audio: Kokoro's isolated single-word renders (what a
# spelled-out character becomes, e.g. "tee") come out both quieter AND less
# articulate than Piper's — gain-boosting the volume didn't fix clarity
# because the underlying render itself is muddier, not just quiet. Piper's
# isolated-word renders measured clean (continuous envelope, near-full-scale
# volume) in the same tests. So EVERY spelled character is synthesized via
# Piper, even on a call that's otherwise using Kokoro for normal speech —
# use the engine that's actually good at this specific job.
async def _spell_char_audio(tok: str, stop_event: "asyncio.Event | None" = None) -> bytes:
    if stop_event is not None and stop_event.is_set():
        return b""
    voice = state.get("piper_voice")
    if voice is None:
        return b""
    from piper.config import SynthesisConfig
    cfg = SynthesisConfig(length_scale=PIPER_SPELL_LENGTH_SCALE)
    loop = asyncio.get_running_loop()

    async def synth():
        gen = voice.synthesize(_spoken_char(tok), cfg)
        while True:
            chunk = await loop.run_in_executor(None, _piper_next, gen)
            if chunk is None:
                break
            yield chunk.audio_int16_bytes

    lock = await _get_tts_lock()
    async with lock:
        if stop_event is not None and stop_event.is_set():
            return b""
        return await _synth_char_bytes(synth(), label=tok, engine="piper")


# ── Piper TTS streaming ──────────────────────────────────────────────────
async def _piper_synth_segments(segments, stop_event: "asyncio.Event | None" = None):
    """Synthesize (is_spelled, content) segments from split_spell_segments():
    plain segments at normal speed, spelled segments one character at a time
    at a slower length-scale with real silence between characters.

    `stop_event`, when set, is checked at every segment/character boundary
    so a barge-in stops the NEXT unit of work immediately — it doesn't wait
    on asyncio task-cancellation semantics, which can be deferred until a
    slow in-flight synth call finishes when a spelled run has many
    characters queued up."""
    voice = state.get("piper_voice")
    if voice is not None:
        # Persistent in-process model — first audio in ~0.1s (vs ~1.8s subprocess)
        from piper.config import SynthesisConfig
        cfg_normal = SynthesisConfig(length_scale=PIPER_LENGTH_SCALE)
        loop = asyncio.get_running_loop()

        async def synth(t, cfg):
            gen = voice.synthesize(t, cfg)
            while True:
                chunk = await loop.run_in_executor(None, _piper_next, gen)
                if chunk is None:
                    break
                yield chunk.audio_int16_bytes

        for is_spell, content in segments:
            if stop_event is not None and stop_event.is_set():
                return
            if not is_spell:
                lock = await _get_tts_lock()
                async with lock:
                    async for b in synth(content, cfg_normal):
                        if stop_event is not None and stop_event.is_set():
                            return
                        yield b
                continue
            for tok in content:
                if stop_event is not None and stop_event.is_set():
                    return
                if tok == SPELL_PAUSE:
                    yield _silence_bytes(SPELL_GROUP_GAP_MS)
                    continue
                audio = await _spell_char_audio(tok, stop_event=stop_event)
                if audio:
                    yield audio
                yield _silence_bytes(SPELL_CHAR_GAP_MS)
        return

    # Fallback: subprocess. Spinning up the Piper CLI per character would be
    # far too slow, so flatten spelled runs into plain, period-separated text.
    flat = flatten_spell_segments(segments)
    proc = await asyncio.create_subprocess_exec(
        "python", "-m", "piper",
        "--model", PIPER_VOICE,
        "--data-dir", PIPER_DATA_DIR,
        "--output-raw",
        "--length-scale", str(PIPER_LENGTH_SCALE),
        stdin=asyncio.subprocess.PIPE,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.DEVNULL,
    )
    try:
        proc.stdin.write(flat.encode())
        await proc.stdin.drain()
        proc.stdin.close()
        while True:
            chunk = await proc.stdout.read(4096)
            if not chunk:
                break
            yield chunk
        await proc.wait()
    finally:
        if proc.returncode is None:
            try:
                proc.kill()
                await proc.wait()
            except ProcessLookupError:
                pass


async def tts_stream(text: str, stop_event: "asyncio.Event | None" = None):
    text = tts_text(text, "piper")
    segments = split_spell_segments(text)
    async for chunk in _piper_synth_segments(segments, stop_event=stop_event):
        yield chunk


# ── Kokoro TTS streaming ────────────────────────────────────────────────
# Serializes in-process Kokoro synthesis (single KPipeline, not thread-safe
# under concurrent calls).
_kokoro_lock = None


async def _get_kokoro_lock():
    global _kokoro_lock
    if _kokoro_lock is None:
        _kokoro_lock = asyncio.Lock()
    return _kokoro_lock


async def _kokoro_synth_segments(segments, stop_event: "asyncio.Event | None" = None):
    from audio import resample
    pipeline = state.get("kokoro_pipeline")
    if not pipeline:
        log.warning("Kokoro not loaded, falling back to Piper")
        async for chunk in _piper_synth_segments(segments, stop_event=stop_event):
            yield chunk
        return
    loop = asyncio.get_running_loop()

    async def synth(t, speed):
        gen = pipeline(t, voice=KOKORO_VOICE, speed=speed)
        while True:
            result = await loop.run_in_executor(None, _piper_next, gen)
            if result is None:
                break
            if result.audio is None or len(result.audio) == 0:
                continue
            audio = result.audio
            if KOKORO_RATE != PIPER_RATE:
                audio = resample(audio, KOKORO_RATE, PIPER_RATE)
            pcm = np.clip(audio * 32767, -32768, 32767).astype(np.int16)
            yield pcm.tobytes()

    try:
        # Stream segment-by-segment (do NOT list() the whole pipeline — that
        # forces the full synthesis before any audio, causing long silence).
        # `stop_event` is checked at every boundary so a barge-in stops the
        # NEXT unit of work immediately, rather than waiting on asyncio
        # task-cancellation timing (see _piper_synth_segments for why that
        # matters on a long spelled run).
        for is_spell, content in segments:
            if stop_event is not None and stop_event.is_set():
                return
            if not is_spell:
                lock = await _get_kokoro_lock()
                async with lock:
                    async for b in synth(content, KOKORO_SPEED):
                        if stop_event is not None and stop_event.is_set():
                            return
                        yield b
                continue
            # Spelled characters go through Piper, not Kokoro — see
            # _spell_char_audio for why.
            for tok in content:
                if stop_event is not None and stop_event.is_set():
                    return
                if tok == SPELL_PAUSE:
                    yield _silence_bytes(SPELL_GROUP_GAP_MS)
                    continue
                audio = await _spell_char_audio(tok, stop_event=stop_event)
                if audio:
                    yield audio
                yield _silence_bytes(SPELL_CHAR_GAP_MS)
    except Exception as e:
        log.error("Kokoro TTS error: %s", e)


async def kokoro_stream(text: str, stop_event: "asyncio.Event | None" = None):
    text = tts_text(text, "kokoro")
    segments = split_spell_segments(text)
    async for chunk in _kokoro_synth_segments(segments, stop_event=stop_event):
        yield chunk


def get_tts_fn(engine: str):
    return {"piper": tts_stream, "kokoro": kokoro_stream}.get(engine, tts_stream)


# ── End-of-call detection ────────────────────────────────────────────────
_END_PHRASES = [
    "bye", "goodbye", "see you", "that's all", "i'm done", "cut the call",
    "end the call", "hang up", "that is all", "no more", "i'm finished",
    "all done", "thank you goodbye",
]


def _is_end_of_call(user_text: str, conversation: list[dict]) -> bool:
    if len(conversation) < 2:
        return False
    lower = user_text.lower().strip()
    if any(p in lower for p in _END_PHRASES):
        return True
    if conversation and any(
        "have a great day" in (m.get("content", "")).lower()
        or "goodbye" in (m.get("content", "")).lower()
        for m in conversation[-3:]
    ):
        if lower in ("okay", "ok", "thanks", "thank you", "bye", "sure", "yes"):
            return True
    return False


# ── Offline STT (full audio, no VAD) ────────────────────────────────────
async def transcribe_full_audio(audio_bytes: bytes) -> str:
    if len(audio_bytes) < 320:
        return ""
    pcm = np.frombuffer(audio_bytes, dtype=np.int16).astype(np.float32) / 32768.0
    async with state["stt_lock"]:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            None, lambda: transcribe_offline(state["stt_model"], pcm)
        )


async def _save_review(session_id: str, conversation: list[dict], full_audio: bytes,
                       call_duration: float, account_uid: str = "", started_at: float = 0.0,
                       tts_engine: str = "", vad_mode: str = "", system_prompt: str = "",
                       stt_total_ms: float = 0.0, stt_count: int = 0,
                       llm_total_ms: float = 0.0, llm_count: int = 0,
                       tts_total_ms: float = 0.0, tts_count: int = 0,
                       peak_prompt_tokens: int = 0, total_completion_tokens: int = 0):
    """Save audio buffer, spawn offline STT, store review data."""
    r = state["redis"]
    audio_key = f"call:{session_id}:audio"
    review_key = f"call:{session_id}:review"

    # Create a call record so browser calls appear in the feed/review list.
    # Use the account context for payer/claim when CALL_RESULT didn't store them.
    existing = await r.hgetall(f"call:{session_id}")
    acct = await r.hgetall(f"account:{account_uid}") if account_uid else {}
    await r.hset(f"call:{session_id}", mapping={
        "claim_id": (existing.get("claim_id") or acct.get("Claim ID")
                     or acct.get("Account Number") or "unknown"),
        "payer": existing.get("payer") or acct.get("Responsible Payer") or "unknown",
        "account_uid": account_uid,
        # Only "completed" if a CALL_RESULT produced a real outcome; otherwise failed
        "status": existing.get("status") or "failed",
        "started_at": str(started_at or time.time()),
        "ended_at": str(time.time()),
        "duration_ms": str(int(call_duration * 1000)),
        "next_action": existing.get("next_action", ""),
        # TTR metrics
        "stt_avg_ms": str(int(stt_total_ms / max(1, stt_count))),
        "llm_avg_ms": str(int(llm_total_ms / max(1, llm_count))),
        "tts_avg_ms": str(int(tts_total_ms / max(1, tts_count))),
        "ttr_avg_ms": str(int((stt_total_ms / max(1, stt_count))
                              + (llm_total_ms / max(1, llm_count))
                              + (tts_total_ms / max(1, tts_count)))),
        "peak_prompt_tokens": str(peak_prompt_tokens),
        "total_completion_tokens": str(total_completion_tokens),
        "context_limit": "4096",
    })

    # Save audio to Redis (24h TTL)
    if full_audio:
        await r.setex(audio_key, 86400, full_audio)

    # Extract real-time user utterances, AI responses, and interleaved transcript
    real_time = []
    ai_responses = []
    transcript = []
    for msg in conversation:
        role = msg.get("role", "")
        content = msg.get("content", "")
        if role == "user":
            cleaned = content.replace("[INSURANCE REP] ", "", 1)
            real_time.append(cleaned)
            transcript.append({"role": "user", "text": cleaned})
        elif role == "assistant":
            ai_responses.append(content)
            transcript.append({"role": "assistant", "text": strip_markers(content)})

    # Store call config/meta + full interleaved transcript
    try:
        await r.set(f"call:{session_id}:transcript", json.dumps(transcript))
        await r.hset(f"call:{session_id}:meta", mapping={
            "stt_model": WHISPER_MODEL_SIZE,
            "tts_engine": tts_engine or TTS_ENGINE,
            "vad_mode": vad_mode or VAD_MODE,
            "llm_model": state["llm_model"],
            "prompt": system_prompt,
            "call_sid": session_id,
        })
    except Exception as e:
        log.error("[%s] Meta save error: %s", session_id, e)

    # Run offline STT in background
    async def _offline_stt():
        try:
            if not full_audio or len(full_audio) < 320:
                review = {"real_time": real_time, "full_audio": "", "ai_responses": ai_responses,
                          "duration_sec": int(call_duration), "audio_size_bytes": len(full_audio)}
                await r.set(review_key, json.dumps(review))
                return
            log.info("[%s] Offline STT on %d bytes...", session_id, len(full_audio))
            transcript = await transcribe_full_audio(full_audio)
            review = {"real_time": real_time, "full_audio": transcript, "ai_responses": ai_responses,
                      "duration_sec": int(call_duration), "audio_size_bytes": len(full_audio)}
            await r.set(review_key, json.dumps(review))
            log.info("[%s] Offline STT done (%d chars)", session_id, len(transcript))
        except Exception as e:
            log.error("[%s] Offline STT error: %s", session_id, e)
            review = {"real_time": real_time, "full_audio": "", "ai_responses": ai_responses,
                      "duration_sec": int(call_duration), "audio_size_bytes": len(full_audio), "error": str(e)}
            await r.set(review_key, json.dumps(review))

    # Keep a strong reference so the task isn't garbage collected mid-run
    task = asyncio.create_task(_offline_stt())
    _bg_tasks.add(task)
    task.add_done_callback(_bg_tasks.discard)


# ── App lifecycle ────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    load_models()
    yield
    await state["redis"].aclose()


app = FastAPI(lifespan=lifespan)


# ════════════════════════════════════════════════════════════════════════
# TELEPHONY  (ported from src/index.ts /voice + /media + /make-call)
# ════════════════════════════════════════════════════════════════════════

def _voice_webhook(payer: str, claim_id: str, account_uid: str, call_sid: str) -> str:
    """Build the Twilio voice webhook URL with properly URL-encoded params."""
    params = urlencode({
        "payer": payer, "claim_id": claim_id,
        "account_uid": account_uid, "local_sid": call_sid,
    })
    return f"{PUBLIC_SCHEME}://{PUBLIC_DOMAIN}/voice?{params}"


@app.post("/make-call")
async def make_call(request: Request):
    """Trigger outbound call via Twilio REST API (ported)."""
    data = await request.json()
    phone = data.get("phone")
    if not phone:
        return JSONResponse({"error": "Phone number is required"}, 400)
    if not all([TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_FROM_NUMBER]):
        return JSONResponse({"error": "Twilio credentials not configured"}, 400)

    # Resolve account context from project_id + row_num if provided
    account_uid = data.get("account_uid", "")
    if data.get("project_id") and data.get("row_num"):
        _uid, acct = await _resolve_account(data["project_id"], int(data["row_num"]))
        if not acct:
            return JSONResponse({"error": f"No account for project {data['project_id']} row {data['row_num']}"}, 404)
        account_uid = _uid or ""
        data.setdefault("payer", acct.get("Responsible Payer", "unknown"))
        data.setdefault("claim_id", acct.get("Claim ID") or acct.get("Account Number") or "unknown")

    payer = data.get("payer", "unknown")
    claim_id = data.get("claim_id", "unknown")
    call_id = f"call-{uuid.uuid4().hex[:12]}"

    # Pre-create the call record keyed by a stable call_id (ported from /voice handler)
    await state["redis"].hset(f"call:{call_id}", mapping={
        "claim_id": claim_id, "payer": payer, "account_uid": account_uid,
        "phone": phone, "status": "dialing", "started_at": str(time.time()),
    })

    from twilio.rest import Client as TwilioClient
    from twilio.base.exceptions import TwilioRestException
    try:
        client = TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        webhook = _voice_webhook(payer, claim_id, account_uid, call_id)
        call = client.calls.create(to=phone, from_=TWILIO_FROM_NUMBER, url=webhook)
        # Keep the stable call_id as the record key; store the Twilio SID as a field
        await state["redis"].hset(f"call:{call_id}", mapping={"twilio_sid": call.sid})
        return {"ok": True, "call_id": call_id, "callSid": call.sid}
    except TwilioRestException as e:
        return JSONResponse({"error": f"Twilio API error: {e.msg}"}, 500)


@app.post("/make-call-enriched")
async def make_call_enriched(request: Request):
    """Trigger outbound call after storing enriched claim data into Redis."""
    data = await request.json()
    phone = data.get("phone")
    claim_id = data.get("claim_id") or data.get("claimId")
    if not phone or not claim_id:
        return JSONResponse({"error": "Phone number and claim_id are required"}, 400)

    account_uid = data.get("account_uid") or data.get("accountUid") or f"claim-{claim_id}"
    payer = data.get("payer") or data.get("provider") or "unknown"
    enriched_data = data.get("enrichedData") or {}

    # Map enriched fields into account dictionary
    record = {
        "UID": account_uid,
        "Claim ID": claim_id,
        "Responsible Payer": payer,
        "Patient Name": enriched_data.get("patientName") or data.get("patientName") or "Unknown Patient",
        "DOS": enriched_data.get("dos") or data.get("dos") or "",
        "Billed Amount": str(enriched_data.get("billedAmount") or data.get("billedAmount") or "0.00"),
        "CPT": enriched_data.get("cpt") or data.get("cpt") or "",
        "Account Number": enriched_data.get("accountNumber") or data.get("accountNumber") or "",
        "Denial Code": enriched_data.get("denialCode") or data.get("denialCode") or "",
        "Call Status": "Pending",
    }

    # Store into Redis hash so CallSession can pick it up
    r = state["redis"]
    await r.hset(f"account:{account_uid}", mapping=record)

    # Call standard make_call logic
    call_id = f"call-{uuid.uuid4().hex[:12]}"
    await r.hset(f"call:{call_id}", mapping={
        "claim_id": claim_id, "payer": payer, "account_uid": account_uid,
        "phone": phone, "status": "dialing", "started_at": str(time.time()),
    })

    if not all([TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_FROM_NUMBER]):
        return JSONResponse({"ok": True, "call_id": call_id, "callSid": call_id, "warning": "Twilio not configured; mock mode active"})

    from twilio.rest import Client as TwilioClient
    from twilio.base.exceptions import TwilioRestException
    try:
        client = TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        webhook = _voice_webhook(payer, claim_id, account_uid, call_id)
        call = client.calls.create(to=phone, from_=TWILIO_FROM_NUMBER, url=webhook)
        await r.hset(f"call:{call_id}", mapping={"twilio_sid": call.sid})
        return {"ok": True, "call_id": call_id, "callSid": call.sid}
    except TwilioRestException as e:
        return JSONResponse({"error": f"Twilio API error: {e.msg}"}, 500)


@app.get("/api/enriched-claim/{claim_id}")
async def get_enriched_claim(claim_id: str):
    """Retrieve enriched claim data stored in Redis."""
    r = state["redis"]
    account_keys = await r.keys("account:*")
    for key in account_keys:
        data = await r.hgetall(key)
        if data.get("Claim ID") == claim_id or data.get("claimId") == claim_id:
            return {"claimId": claim_id, "data": data}
    return JSONResponse({"error": "Claim not found"}, 404)




@app.api_route("/voice", methods=["GET", "POST"])
async def voice_webhook(request: Request):
    """Twilio webhook → TwiML with Media Stream (ported)."""
    q = request.query_params
    form = await request.form() if request.method == "POST" else {}
    # Prefer local_sid (our call_id) so the media stream + live transcript key on
    # the id returned by /make-call — the one the UI polls.
    call_sid = q.get("local_sid") or q.get("CallSid") or (form.get("CallSid") if form else None) or "unknown"

    # Ensure call record exists with context
    existing = await state["redis"].hgetall(f"call:{call_sid}")
    if not existing:
        await state["redis"].hset(f"call:{call_sid}", mapping={
            "claim_id": q.get("claim_id", "unknown"),
            "payer": q.get("payer", "unknown"),
            "account_uid": q.get("account_uid", ""),
            "status": "dialing", "started_at": str(time.time()),
        })

    twiml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
  <Connect>
    <Stream url="wss://{PUBLIC_DOMAIN}/media/{call_sid}">
      <Parameter name="payer" value="{q.get('payer', 'unknown')}"/>
      <Parameter name="claim_id" value="{q.get('claim_id', 'unknown')}"/>
      <Parameter name="account_uid" value="{q.get('account_uid', '')}"/>
    </Stream>
  </Connect>
</Response>"""
    return Response(content=twiml, media_type="text/xml")


@app.websocket("/media/{call_sid}")
async def media_stream(ws: WebSocket, call_sid: str):
    """Twilio Media Streams WebSocket → CallSession (ported from DO)."""
    _cfg = await _get_config()
    engine = _cfg["tts_engine"]
    call_record = await state["redis"].hgetall(f"call:{call_sid}")
    if call_record and call_record.get("tts_engine"):
        engine = call_record["tts_engine"]
    session = CallSession(ws, call_sid, {
        "redis": state["redis"],
        "stt_model": state["stt_model"],
        "stt_lock": state["stt_lock"],
        "llm_client": state["llm_client"],
        "llm_model": state["llm_model"],
        "tts_stream_fn": get_tts_fn(engine),
        "tts_engine": engine,
        "vad_mode": _cfg["vad_mode"],
        "stt_model_name": WHISPER_MODEL_SIZE,
        "use_silero": _cfg["vad_mode"] == "silero",
        "opik": state.get("opik"),
    })
    await session.run()


@app.post("/call-result")
async def call_result(request: Request):
    """External call completion hook (ported)."""
    data = await request.json()
    call_sid = data.get("callSid")
    if not call_sid:
        return JSONResponse({"error": "callSid required"}, 400)
    await state["redis"].hset(f"call:{call_sid}", mapping={
        **{k: str(v) for k, v in data.items() if k != "callSid"},
        "ended_at": str(time.time()), "status": "completed",
    })
    await state["redis"].publish("call-updates", json.dumps(data))
    return {"ok": True}


@app.post("/retry/{call_sid}")
async def retry_call(call_sid: str):
    """Re-queue a failed call (ported)."""
    data = await state["redis"].hgetall(f"call:{call_sid}")
    if not data or data.get("status") != "failed":
        return JSONResponse({"error": "Not found or not failed"}, 404)
    await state["redis"].hset(f"call:{call_sid}", mapping={
        "status": "queued",
        "retry_count": str(int(data.get("retry_count", 0)) + 1),
        "started_at": str(time.time()),
    })
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════
# DATA APIs  (ported: calls list, CSV export, secrets check)
# ════════════════════════════════════════════════════════════════════════

_SUFFIX_EXCLUDE = (":audio", ":review", ":transcript", ":meta", ":live")


def _call_record_keys(keys):
    """Filter Redis keys to actual call record hashes (exclude sub-keys)."""
    return [k for k in keys if not k.endswith(_SUFFIX_EXCLUDE)]


# ── Excel header mapping ─────────────────────────────────────────────────
HEADER_ALIASES = {
    # objective
    "final comments": "AR Final Comments",
    "ar final comments": "AR Final Comments",
    # patient / demographics
    "patient": "Patient Name",
    "patient name": "Patient Name",
    "patient's name": "Patient Name",
    "member id": "Member ID",
    "member": "Member ID",
    "dob": "DOB",
    "date of birth": "DOB",
    "dos": "DOS",
    "date of service": "DOS",
    "cpt": "CPT",
    "cpt code": "CPT",
    "billed amount": "Billed Amount",
    "billed": "Billed Amount",
    "amount billed": "Billed Amount",
    # payer / billing
    "responsible party": "Responsible Payer",
    "responsible payer": "Responsible Payer",
    "payer": "Responsible Payer",
    "account number": "Account Number",
    "account #": "Account Number",
    "claim id": "Claim ID",
    "claim number": "Claim ID",
    # group / provider
    "group": "Group",
    "group name": "Group Name",
    "group npi": "Group NPI",
    "tax id": "Tax ID",
    "billing/provider address": "Billing/Provider Address",
    "billing address": "Billing/Provider Address",
    "provider address": "Billing/Provider Address",
    "pay-to-address": "Pay-to-address",
    "pay to address": "Pay-to-address",
    "individual npi": "Individual NPI",
    # workflow / outcome
    "call status": "Call Status",
    "status": "Call Status",
    "denial code": "Denial Code",
    "denial": "Denial Code",
    "notes": "Notes",
    "uid": "UID",
}


def _normalize_header(h: str) -> str:
    """Map a source Excel header to the canonical field name (case-insensitive)."""
    key = h.strip().lower()
    return HEADER_ALIASES.get(key, h.strip())


@app.get("/api/stats")
async def stats():
    """True aggregate counts (not capped to the recent-20 list)."""
    keys = _call_record_keys(await state["redis"].keys("call:*"))
    total = len(keys)
    completed = 0
    total_dur = 0
    for key in keys:
        d = await state["redis"].hgetall(key)
        if d.get("status") == "completed":
            completed += 1
        total_dur += int(d.get("duration_ms") or 0)
    projects = len(await state["redis"].keys("project:*:rows"))
    return {
        "total_calls": total,
        "completed": completed,
        "projects": projects,
        "total_duration_ms": total_dur,
    }


@app.get("/api/calls")
async def list_calls(page: int | None = None, limit: int = 15):
    r = state["redis"]
    keys = _call_record_keys(await r.keys("call:*"))
    calls = []
    for key in keys:
        data = await r.hgetall(key)
        if data:
            calls.append({"callSid": key.replace("call:", ""), **data})
    calls.sort(key=lambda c: float(c.get("started_at", 0)), reverse=True)
    if page is None:
        return calls[:20]  # backward-compat for legacy /dashboard
    limit = max(1, min(limit, 100))
    page = max(1, page)
    start = (page - 1) * limit
    total = len(calls)
    return {
        "calls": calls[start:start + limit],
        "total": total,
        "page": page,
        "limit": limit,
        "pages": max(1, -(-total // limit)),
    }


@app.get("/export.csv")
async def export_csv():
    r = state["redis"]
    keys = _call_record_keys(await r.keys("call:*"))
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["call_sid", "timestamp", "payer", "claim_id", "status",
                     "amount", "next_action", "duration_sec"])
    for key in keys:
        d = await r.hgetall(key)
        if d.get("status") in ("completed", "failed"):
            ts = float(d.get("ended_at") or d.get("started_at") or 0)
            writer.writerow([
                key.replace("call:", ""),
                time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime(ts)),
                d.get("payer", ""), d.get("claim_id", ""), d.get("status", ""),
                d.get("amount", ""), d.get("next_action", ""),
                round(float(d.get("duration_ms", 0)) / 1000),
            ])
    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition":
                 f'attachment; filename="ar-calls-{time.strftime("%Y-%m-%d")}.csv"'},
    )


@app.get("/api/check-secrets")
async def check_secrets():
    return {
        "twilio_sid": bool(TWILIO_ACCOUNT_SID),
        "twilio_token": bool(TWILIO_AUTH_TOKEN),
        "twilio_from": bool(TWILIO_FROM_NUMBER),
        "vllm_url": VLLM_BASE_URL,
        "whisper_model": WHISPER_MODEL_SIZE,
        "llm_model": LLM_MODEL,
    }


# ════════════════════════════════════════════════════════════════════════
# EXCEL  (ported from src/index.ts upload/accounts/export)
# ════════════════════════════════════════════════════════════════════════

async def _resolve_account(project_id: str, row_num: int) -> tuple[str | None, dict | None]:
    """Resolve (project_id, row_num) → (account_uid, account_record). Row 1 = first data row."""
    raw = await state["redis"].get(f"project:{project_id}:rows")
    if not raw:
        return None, None
    uids = json.loads(raw)
    if not (1 <= row_num <= len(uids)):
        return None, None
    uid = uids[row_num - 1]
    acct = await state["redis"].hgetall(f"account:{uid}") or None
    return uid, acct


async def _effective_prompt(account: dict | None) -> str:
    """System prompt for an account: custom override if set, else the built default.
    Notes are appended in either case."""
    if not account:
        return build_call_prompt("GREETING", None, None, None)
    uid = account.get("UID", "")
    custom = None
    if uid:
        custom = await state["redis"].get(f"account:{uid}:llm_context")
    notes = account.get("Notes", "")
    if custom:
        prompt = custom
        if notes:
            prompt += f"\n\n[PRIOR CALL NOTES]\n{notes}"
        return prompt
    return build_call_prompt("GREETING", None, None, account)


@app.get("/api/accounts/{account_uid}/llm-context")
async def get_llm_context(account_uid: str):
    account = await state["redis"].hgetall(f"account:{account_uid}") or None
    custom = await state["redis"].get(f"account:{account_uid}:llm_context") or ""
    original = build_call_prompt("GREETING", None, None, account)
    effective = await _effective_prompt(account)
    return {"account_uid": account_uid, "custom": custom, "original": original, "effective": effective}


@app.post("/api/accounts/{account_uid}/llm-context")
async def update_llm_context(account_uid: str, request: Request):
    data = await request.json()
    key = f"account:{account_uid}:llm_context"
    if data.get("reset"):
        await state["redis"].delete(key)
    else:
        ctx = (data.get("context") or "").strip()
        if not ctx:
            return JSONResponse({"error": "context is required"}, 400)
        await state["redis"].set(key, ctx)
    account = await state["redis"].hgetall(f"account:{account_uid}") or None
    return {"ok": True, "custom": (await state["redis"].get(key)) or "",
            "effective": await _effective_prompt(account)}


@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return JSONResponse({"error": "Sheet is empty"}, 400)

        headers = [_normalize_header(str(h) if h else f"col_{i}") for i, h in enumerate(rows[0])]
        r = state["redis"]

        # Generate a unique project id for this upload
        project_id = f"proj-{uuid.uuid4().hex[:10]}"

        uids = []
        for i, row in enumerate(rows[1:]):
            record = {}
            for j, val in enumerate(row):
                if j < len(headers) and val is not None:
                    record[headers[j]] = val.isoformat() if hasattr(val, "isoformat") else str(val)
            # "Final Comments" → objective (AR Final Comments) is handled by HEADER_ALIASES
            uid = record.get("UID") or f"{project_id}-{i}-{int(time.time())}"
            record["UID"] = uid
            record.setdefault("Call Status", "Pending")
            uids.append(uid)
            await r.hset(f"account:{uid}", mapping=record)

        # Project-scoped row index: index 0 = row 1 = first data row
        await r.set(f"project:{project_id}:rows", json.dumps(uids))
        await r.set(f"project:{project_id}:headers", json.dumps(headers))
        # Keep legacy lists for the UI (last project shown)
        await r.set("accounts-headers", json.dumps(headers))
        await r.set("accounts-list", json.dumps(uids))
        return {"ok": True, "count": len(uids), "project_id": project_id}
    except Exception as e:
        log.error("Excel parse error: %s", e)
        return JSONResponse({"error": f"Parsing error: {e}"}, 500)


@app.get("/api/projects/{project_id}/accounts/{row_num}")
async def get_project_account(project_id: str, row_num: int):
    """Resolve an account by project + row number (row 1 = first data row)."""
    uid, acct = await _resolve_account(project_id, row_num)
    if not acct:
        return JSONResponse({"error": f"No account for project {project_id} row {row_num}"}, 404)
    return {"account_uid": uid, "account": acct}


@app.get("/api/projects")
async def list_projects():
    """List all uploaded projects with row counts."""
    keys = await state["redis"].keys("project:*:rows")
    projects = []
    for key in keys:
        pid = key.replace("project:", "").replace(":rows", "")
        raw = await state["redis"].get(key)
        try:
            count = len(json.loads(raw)) if raw else 0
        except json.JSONDecodeError:
            count = 0
        projects.append({"project_id": pid, "rows": count})
    projects.sort(key=lambda p: p["project_id"], reverse=True)
    return projects


@app.delete("/api/projects/{project_id}")
async def delete_project(project_id: str):
    """Delete a project: its metadata, accounts, and the call records they own."""
    r = state["redis"]
    raw = await r.get(f"project:{project_id}:rows")
    if not raw:
        return JSONResponse({"error": f"Project {project_id} not found"}, 404)
    uids = json.loads(raw)
    uid_set = set(uids)

    await r.delete(f"project:{project_id}:rows", f"project:{project_id}:headers")

    for uid in uids:
        await r.delete(f"account:{uid}")
        await r.delete(f"account:{uid}:llm_context")
        await r.delete(f"account:{uid}:active")
        await r.delete(f"account:{uid}:active:meta")

    # Remove call records (and sub-keys) that belong to this project's accounts
    call_keys = await r.keys("call:*")
    base_keys = _call_record_keys(call_keys)
    deleted_calls = 0
    for key in base_keys:
        data = await r.hgetall(key)
        if data.get("account_uid") in uid_set:
            base = key.replace("call:", "")
            await r.delete(key)
            for suffix in (":audio", ":review", ":transcript", ":meta", ":live"):
                await r.delete(f"call:{base}{suffix}")
            deleted_calls += 1

    return {"ok": True, "deleted_accounts": len(uids), "deleted_calls": deleted_calls}


@app.get("/api/projects/{project_id}/accounts")
async def list_project_accounts(project_id: str):
    raw = await state["redis"].get(f"project:{project_id}:rows")
    if not raw:
        return []
    uids = json.loads(raw)
    accounts = []
    for uid in uids:
        row = await state["redis"].hgetall(f"account:{uid}")
        if row:
            accounts.append(row)
    return accounts


@app.post("/api/notes")
async def add_note(request: Request):
    """Add a note to an account (used as context on the next call).
    Body: {account_uid} or {project_id, row_num} plus {note}."""
    data = await request.json()
    note = (data.get("note") or "").strip()
    if not note:
        return JSONResponse({"error": "note is required"}, 400)

    uid = data.get("account_uid", "")
    if not uid and data.get("project_id") and data.get("row_num"):
        uid, _ = await _resolve_account(data["project_id"], int(data["row_num"]))
    if not uid:
        return JSONResponse({"error": "account_uid or project_id+row_num required"}, 400)

    account = await state["redis"].hgetall(f"account:{uid}") or {}
    prior = account.get("Notes", "")
    now = time.strftime("%Y-%m-%d %H:%M")
    merged = (prior + "\n" if prior else "") + f"[{now}] {note}"
    await state["redis"].hset(f"account:{uid}", mapping={"Notes": merged})
    return {"ok": True, "account_uid": uid, "notes": merged}


async def _archive_chat(account_uid: str) -> tuple[str | None, dict | None]:
    """Archive the active chat into an ended session (with config + timing)."""
    chat_key = f"chat:{account_uid}:active"
    raw = await state["redis"].lrange(chat_key, 0, -1)
    if not raw:
        return None, None
    turns = [json.loads(x) for x in raw]
    m = await state["redis"].hgetall(f"chat:{account_uid}:active:meta")
    sid = f"chat-{uuid.uuid4().hex[:12]}"
    turns_key = f"chat:{account_uid}:session:{sid}"
    for t in turns:
        await state["redis"].rpush(turns_key, json.dumps(t))

    llm_count = int(m.get("llm_count", 0))
    llm_avg = int(float(m.get("llm_total_ms", 0)) / max(1, llm_count))
    meta = {
        "account_uid": account_uid,
        "started_at": turns[0].get("ts", time.time()),
        "ended_at": time.time(),
        "count": len(turns),
        "preview": (turns[0].get("text") or "")[:120],
        "llm_model": m.get("llm_model", ""),
        "prompt": m.get("prompt", ""),
        "llm_avg_ms": llm_avg,
        "llm_count": llm_count,
    }
    await state["redis"].set(f"chat:session:{sid}", json.dumps(meta))
    await state["redis"].rpush(f"chat:{account_uid}:history", sid)
    await state["redis"].delete(chat_key)
    await state["redis"].delete(f"chat:{account_uid}:active:meta")
    return sid, meta


# ── Direct chat with the LLM (bypasses STT/TTS, isolated from calls/review) ─
@app.post("/api/chat")
async def chat_with_llm(request: Request):
    data = await request.json()
    message = (data.get("message") or "").strip()
    if not message:
        return JSONResponse({"error": "message required"}, 400)

    account_uid = data.get("account_uid", "")
    if not account_uid and data.get("project_id") and data.get("row_num"):
        account_uid, _ = await _resolve_account(data["project_id"], int(data["row_num"]))
    if not account_uid:
        return JSONResponse({"error": "account_uid or project_id+row_num required"}, 400)

    account = await state["redis"].hgetall(f"account:{account_uid}") or None
    system_prompt = await _effective_prompt(account)
    chat_key = f"chat:{account_uid}:active"
    meta_key = f"chat:{account_uid}:active:meta"

    raw = await state["redis"].lrange(chat_key, 0, -1)
    history = [json.loads(x) for x in raw] if raw else []

    messages = [{"role": "system", "content": system_prompt}]
    for t in history[-16:]:
        messages.append({"role": t["role"], "content": t["text"]})
    messages.append({"role": "user", "content": message})

    trace = opik_start_trace("chat.turn", f"chat:{account_uid}",
                             {"messages": messages[-2:]},
                             {"type": "chat", "account_uid": account_uid})

    llm_t0 = time.time()
    try:
        resp = await state["llm_client"].chat.completions.create(
            model=state["llm_model"], messages=messages, max_tokens=300, temperature=0)
        reply_raw = resp.choices[0].message.content.strip()
    except Exception as e:
        opik_end_trace(trace, error=str(e))
        return JSONResponse({"error": f"LLM error: {e}"}, 500)
    llm_ms = (time.time() - llm_t0) * 1000

    _chat_usage = getattr(resp, "usage", None)
    opik_span(trace, "chat.llm", "llm", {"messages": messages[-2:]}, reply_raw,
              llm_t0, time.time(), model=state["llm_model"],
              metadata={"usage": {"prompt_tokens": getattr(_chat_usage, "prompt_tokens", None),
                                  "completion_tokens": getattr(_chat_usage, "completion_tokens", None),
                                  "total_tokens": getattr(_chat_usage, "total_tokens", None)}} if _chat_usage else None)
    opik_end_trace(trace, output=reply_raw)

    # Track chat timing/config in the ACTIVE meta (never touches call records)
    await state["redis"].hincrbyfloat(meta_key, "llm_total_ms", llm_ms)
    await state["redis"].hincrby(meta_key, "llm_count", 1)
    await state["redis"].hset(meta_key, mapping={
        "llm_model": state["llm_model"], "prompt": system_prompt, "account_uid": account_uid,
    })

    reply = strip_markers(reply_raw)

    # Detect end-of-chat: CALL_RESULT marker, empty reply, or a farewell phrase
    has_call_result = bool(re.search(r"CALL_RESULT", reply_raw, re.IGNORECASE))
    is_empty = not reply.strip()
    is_farewell = any(p in reply.lower() for p in
                      ["goodbye", "good bye", "ending the chat", "end the chat",
                       "that concludes", "no further questions", "have a great day"])
    should_end = has_call_result or is_empty or is_farewell

    now = time.time()
    await state["redis"].rpush(chat_key, json.dumps({"role": "user", "text": message, "ts": now}))
    if reply:
        await state["redis"].rpush(chat_key, json.dumps({"role": "assistant", "text": reply, "ts": now}))
    if len(history) > 200:
        await state["redis"].ltrim(chat_key, -200, -1)

    # Auto-archive to history when the AI ends the chat
    if should_end:
        sid, meta = await _archive_chat(account_uid)
        return {"ok": True, "reply": reply, "ended": True, "session_id": sid}

    return {"ok": True, "reply": reply, "ended": False}


@app.get("/api/chat/{account_uid}")
async def get_chat(account_uid: str):
    raw = await state["redis"].lrange(f"chat:{account_uid}:active", 0, -1)
    return [json.loads(x) for x in raw] if raw else []


@app.post("/api/chat/{account_uid}/end")
async def end_chat(account_uid: str):
    """End the active chat → archive it to history and clear the active chat."""
    sid, meta = await _archive_chat(account_uid)
    if not sid:
        return {"ok": True, "ended": False}
    return {"ok": True, "ended": True, "session_id": sid, "meta": meta}


@app.get("/api/chat/{account_uid}/history")
async def chat_history(account_uid: str):
    sids = await state["redis"].lrange(f"chat:{account_uid}:history", 0, -1)
    out = []
    for sid in sids:
        raw = await state["redis"].get(f"chat:session:{sid}")
        if raw:
            try:
                m = json.loads(raw)
                m["session_id"] = sid
                out.append(m)
            except json.JSONDecodeError:
                pass
    out.sort(key=lambda m: m.get("ended_at", 0), reverse=True)
    return out


@app.get("/api/chat/session/{session_id}")
async def chat_session_detail(session_id: str):
    raw_meta = await state["redis"].get(f"chat:session:{session_id}")
    if not raw_meta:
        return JSONResponse({"error": "Chat session not found"}, 404)
    meta = json.loads(raw_meta)
    uid = meta.get("account_uid", "")
    raw_turns = await state["redis"].lrange(f"chat:{uid}:session:{session_id}", 0, -1)
    turns = [json.loads(x) for x in raw_turns] if raw_turns else []
    account = await state["redis"].hgetall(f"account:{uid}") or {}
    prompt = meta.get("prompt", "")
    if not prompt:
        # Back-fill for sessions created before prompt tracking existed
        prompt = await _effective_prompt(account)
    return {
        "session_id": session_id,
        "meta": meta,
        "turns": turns,
        "config": {"llm_model": meta.get("llm_model", ""), "prompt": prompt},
        "timing": {
            "llm_avg_ms": meta.get("llm_avg_ms", 0),
            "llm_count": meta.get("llm_count", 0),
            "ttr_avg_ms": meta.get("llm_avg_ms", 0),
        },
        "account": {
            "Patient Name": account.get("Patient Name", ""),
            "Responsible Payer": account.get("Responsible Payer", ""),
            "DOS": account.get("DOS", ""),
            "Claim ID": account.get("Claim ID") or account.get("Account Number", ""),
        },
    }


@app.get("/api/accounts/{account_uid}/calls")
async def account_call_history(account_uid: str):
    """Call history for a given account/claim."""
    keys = _call_record_keys(await state["redis"].keys("call:*"))
    calls = []
    for key in keys:
        data = await state["redis"].hgetall(key)
        if data and data.get("account_uid") == account_uid:
            calls.append({"call_id": key.replace("call:", ""), **data})
    calls.sort(key=lambda c: float(c.get("started_at", 0)), reverse=True)
    return calls


@app.get("/api/accounts")
async def list_accounts():
    r = state["redis"]
    raw = await r.get("accounts-list")
    if not raw:
        return []
    accounts = []
    for uid in json.loads(raw):
        row = await r.hgetall(f"account:{uid}")
        if row:
            accounts.append(row)
    return accounts


@app.get("/api/export-excel")
async def export_excel():
    r = state["redis"]
    raw = await r.get("accounts-list")
    if not raw:
        return Response("No accounts to export", 404)
    uids = json.loads(raw)
    headers_raw = await r.get("accounts-headers")
    headers = json.loads(headers_raw) if headers_raw else []

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for uid in uids:
        row = await r.hgetall(f"account:{uid}")
        ws.append([row.get(h, "") for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Calling_Accounts_Updated.xlsx"'},
    )


# ════════════════════════════════════════════════════════════════════════
# HEALTH + DASHBOARD
# ════════════════════════════════════════════════════════════════════════

@app.get("/health")
async def health():
    return {"status": "ok", "whisper": WHISPER_MODEL_SIZE, "llm": state["llm_model"]}


@app.get("/api/current-llm")
async def current_llm():
    return {
        "model": state["llm_model"],
        "options": state["llm_options"],
        "switching": state.get("llm_switching", False),
    }


@app.post("/api/switch-llm")
async def switch_llm(request: Request):
    data = await request.json()
    model_name = data.get("model", "")
    # Find the full model ID from options
    model_id = None
    for label, mid in state["llm_options"].items():
        if model_name in (label, mid):
            model_id = mid
            break
    if not model_id:
        return JSONResponse({"error": f"Unknown model: {model_name}"}, 400)
    if model_id == state["llm_model"]:
        return {"ok": True, "model": model_id, "switching": False}

    state["llm_switching"] = True
    # Write to .env on host
    try:
        import subprocess
        env_path = "/project/.env"
        with open(env_path) as f:
            lines = f.readlines()
        found = False
        for i, line in enumerate(lines):
            if line.startswith("LLM_MODEL="):
                lines[i] = f"LLM_MODEL={model_id}\n"
                found = True
                break
        if not found:
            lines.append(f"LLM_MODEL={model_id}\n")
        with open(env_path, "w") as f:
            f.writelines(lines)
    except Exception as e:
        log.warning("Could not update .env: %s", e)

    # Update in-memory model
    old_model = state["llm_model"]
    state["llm_model"] = model_id

    # Restart vLLM in background
    async def _restart_vllm():
        try:
            import subprocess
            subprocess.run(["docker", "restart", "vllm"], timeout=120)
            log.info("vLLM restarted with model %s", model_id)
        except Exception as e:
            log.error("vLLM restart failed: %s", e)
            state["llm_model"] = old_model
        finally:
            state["llm_switching"] = False

    asyncio.create_task(_restart_vllm())
    return {"ok": True, "model": model_id, "switching": True}


@app.get("/api/calls/{session_id}/live")
async def get_live_transcript(session_id: str):
    """Live transcript for an in-progress call (UI polls this)."""
    raw = await state["redis"].lrange(f"call:{session_id}:live", 0, -1)
    turns = [json.loads(x) for x in raw] if raw else []
    record = await state["redis"].hgetall(f"call:{session_id}")
    active = bool(record) and "ended_at" not in record
    return {"call_id": session_id, "active": active, "transcript": turns}


@app.get("/api/calls/{session_id}/review")
async def get_review(session_id: str):
    review_key = f"call:{session_id}:review"
    data = await state["redis"].get(review_key)
    if not data:
        return JSONResponse({"error": "Review not found"}, 404)
    return JSONResponse(json.loads(data))


@app.get("/api/calls/{call_id}/detail")
async def get_call_detail(call_id: str):
    """Full detail for a call: config (STT/TTS/LLM/prompt), interleaved transcript, review."""
    r = state["redis"]
    record = await r.hgetall(f"call:{call_id}")
    if not record:
        return JSONResponse({"error": "Call not found"}, 404)

    meta = await r.hgetall(f"call:{call_id}:meta")

    transcript = []
    raw_t = await r.get(f"call:{call_id}:transcript")
    if raw_t:
        try:
            transcript = json.loads(raw_t)
        except json.JSONDecodeError:
            transcript = []

    review = None
    raw_r = await r.get(f"call:{call_id}:review")
    if raw_r:
        try:
            review = json.loads(raw_r)
        except json.JSONDecodeError:
            review = None

    # Old calls (pre-transcript feature) — reconstruct interleaved transcript
    # from the review arrays if no stored transcript exists.
    if not transcript and review:
        rt = review.get("real_time", []) or []
        ai = review.get("ai_responses", []) or []
        if ai:
            transcript.append({"role": "assistant", "text": ai[0]})
        for i, u in enumerate(rt):
            transcript.append({"role": "user", "text": u})
            if i + 1 < len(ai):
                transcript.append({"role": "assistant", "text": ai[i + 1]})
        if transcript:
            try:
                await r.set(f"call:{call_id}:transcript", json.dumps(transcript))
            except Exception:
                pass

    return {
        "call_id": call_id,
        "twilio_sid": record.get("twilio_sid", ""),
        "call": {k: v for k, v in record.items()},
        "config": {
            "stt_model": meta.get("stt_model", "unknown"),
            "tts_engine": meta.get("tts_engine", "unknown"),
            "vad_mode": meta.get("vad_mode", "unknown"),
            "llm_model": meta.get("llm_model", "unknown"),
        },
        "prompt": meta.get("prompt", ""),
        "transcript": transcript,
        "review": review,
    }


# ── App config (TTS/VAD defaults, persisted in Redis, editable via API) ─
CONFIG_KEY = "config:app"


async def _get_config() -> dict:
    cfg = await state["redis"].hgetall(CONFIG_KEY)
    return {
        "tts_engine": cfg.get("tts_engine", TTS_ENGINE),
        "vad_mode": cfg.get("vad_mode", VAD_MODE),
        "stay_awake": cfg.get("stay_awake", "0"),
    }


async def _set_config(**kwargs):
    await state["redis"].hset(CONFIG_KEY, mapping={k: str(v) for k, v in kwargs.items()})


@app.get("/api/config")
async def get_config():
    cfg = await _get_config()
    return {
        "tts_engine": cfg["tts_engine"],
        "vad_mode": cfg["vad_mode"],
        "stay_awake": cfg["stay_awake"] == "1",
        "tts_options": ["piper", "kokoro"],
        "vad_options": ["silero", "rms"],
        "llm_model": state["llm_model"],
        "llm_options": state["llm_options"],
    }


@app.post("/api/config")
async def post_config(request: Request):
    data = await request.json()
    updates = {}
    if "tts_engine" in data:
        if data["tts_engine"] not in ("piper", "kokoro"):
            return JSONResponse({"error": "tts_engine must be piper or kokoro"}, 400)
        updates["tts_engine"] = data["tts_engine"]
    if "vad_mode" in data:
        if data["vad_mode"] not in ("silero", "rms"):
            return JSONResponse({"error": "vad_mode must be silero or rms"}, 400)
        updates["vad_mode"] = data["vad_mode"]
    if "stay_awake" in data:
        updates["stay_awake"] = "1" if data["stay_awake"] else "0"
    if updates:
        await _set_config(**updates)
    return {"ok": True, "config": await _get_config()}


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard():
    try:
        from pathlib import Path
        path = Path(__file__).parent / "static" / "dashboard.html"
        if path.exists():
            return HTMLResponse(path.read_text())
    except Exception:
        pass
    return DASHBOARD_HTML


@app.get("/test", response_class=HTMLResponse)
async def voice_test():
    return """<!DOCTYPE html>
<html><head><title>Voice Test</title></head>
<body style="font-family:sans-serif;max-width:600px;margin:2rem auto;">
<h2>Voice Agent Test</h2>
<button id="toggle" onclick="toggle()" style="padding:1rem 2rem;font-size:1.2rem;background:#2563eb;color:white;border:none;border-radius:8px;cursor:pointer">Start Test</button>
<p id="status" style="margin:1rem 0;font-weight:bold;">Disconnected</p>
<div id="log" style="height:400px;overflow-y:auto;background:#1a1a1a;color:#0f0;padding:1rem;border-radius:8px;font-family:monospace;"></div>
<script>
let ws, stream, ctx, src, proc, playCtx, audioQ = [], playing = false, currentSrc = null;
const logDiv = document.getElementById('log'), btn = document.getElementById('toggle'), st = document.getElementById('status');
let audioRate = 22050;
function lg(m,c){const d=document.createElement('div');d.textContent=m;d.style.color=c||'#888';logDiv.appendChild(d);logDiv.scrollTop=logDiv.scrollHeight}
function toggle(){if(ws&&ws.readyState===1){ws.close();return}connect()}
async function connect(){
  playCtx = new AudioContext();
  ws=new WebSocket('wss://'+location.host+'/ws/test_'+Math.random().toString(36).slice(2));
  ws.binaryType='arraybuffer';
  btn.disabled=true;btn.textContent='Connecting...';st.textContent='Connecting';
  ws.onopen=()=>{btn.textContent='Stop';st.textContent='Connected';startMic()};
  ws.onclose=()=>{btn.textContent='Start Test';st.textContent='Disconnected';stopMic();ws=null};
  ws.onmessage=e=>{
    if(typeof e.data==='string'){const m=JSON.parse(e.data);
      if(m.type==='config')audioRate=m.sample_rate;
      else lg(m.text,m.type==='transcript'?'#8cf':'#fc8');
    }else{
      const v=new Uint8Array(e.data);
      if(v[0]===1){audioQ.push(v.slice(1));if(!playing)playNext();}
      else if(v[0]===2){audioQ=[];if(playing&&currentSrc){playing=false;try{currentSrc.stop()}catch(e){}}}
    }
  };
}
function playNext(){
  if(!audioQ.length||!playCtx){playing=false;return}
  playing=true;
  const total=audioQ.reduce((s,c)=>s+c.length,0);
  const pcm=new Int16Array(total/2);let off=0;
  while(audioQ.length){const c=audioQ.shift();pcm.set(new Int16Array(c.buffer,c.byteOffset,c.length/2),off);off+=c.length/2}
  const buf=playCtx.createBuffer(1,pcm.length,audioRate);
  const ch=buf.getChannelData(0);
  for(let i=0;i<pcm.length;i++)ch[i]=pcm[i]/32768;
  const s=playCtx.createBufferSource();currentSrc=s;
  s.buffer=buf;s.connect(playCtx.destination);
  s.onended=()=>{playing=false;currentSrc=null;if(audioQ.length)playNext()};
  s.start();
}
async function startMic(){
  ctx=new AudioContext({sampleRate:16000});
  stream=await navigator.mediaDevices.getUserMedia({audio:true});
  src=ctx.createMediaStreamSource(stream);
  proc=ctx.createScriptProcessor(4096,1,1);
  proc.onaudioprocess=e=>{if(!ws||ws.readyState!==1)return;const inp=e.inputBuffer.getChannelData(0);const b=new Int16Array(inp.length);for(let i=0;i<inp.length;i++)b[i]=Math.max(-32768,Math.min(32767,inp[i]*32768));ws.send(b.buffer)};
  src.connect(proc);proc.connect(ctx.destination);lg('Mic started','#4c4')
}
function stopMic(){if(proc){proc.disconnect();proc=null}if(src){src.disconnect();src=null}if(stream){stream.getTracks().forEach(t=>t.stop());stream=null}if(ctx){ctx.close();ctx=null}}
</script></body></html>"""


@app.websocket("/ws/{session_id}")
async def browser_voice_loop(ws: WebSocket, session_id: str):
    """Direct browser voice test (16kHz PCM) with barge-in."""
    await ws.accept()
    await ws.send_json({"type": "config", "sample_rate": PIPER_RATE})

    # Parse query params from WebSocket URL (override app-config defaults)
    account_uid = ""
    project_id = ""
    row_num = ""
    _cfg = await _get_config()
    tts_engine = _cfg["tts_engine"]
    vad_mode = _cfg["vad_mode"]
    try:
        qs = str(ws.url).split("?")[1] if "?" in str(ws.url) else ""
        for part in qs.split("&"):
            k, _, v = part.partition("=")
            if k == "account_uid" and v:
                account_uid = v
            elif k == "project_id" and v:
                project_id = v
            elif k == "row_num" and v:
                row_num = v
            elif k == "tts" and v:
                tts_engine = v
            elif k == "vad" and v:
                vad_mode = v
    except Exception:
        pass

    # Resolve account context from project_id + row_num if provided
    if project_id and row_num:
        _uid, _acct = await _resolve_account(project_id, int(row_num))
        if _acct:
            account_uid = _uid or ""
        else:
            log.warning("[%s] No account for project %s row %s", session_id, project_id, row_num)

    tts_fn = get_tts_fn(tts_engine)
    use_silero = vad_mode == "silero"

    account = None
    if account_uid:
        account = await state["redis"].hgetall(f"account:{account_uid}") or None

    system_prompt = await _effective_prompt(account)
    conversation: list[dict] = [{"role": "system", "content": system_prompt}]

    # TTR timing accumulators (ms) — must exist before greeting TTS
    _tts_times = []
    stt_total_ms = 0.0
    stt_count = 0
    llm_total_ms = 0.0
    llm_count = 0
    peak_prompt_tokens = 0
    total_completion_tokens = 0
    tts_total_ms = 0.0
    tts_count = 0

    tts_task = None
    # Set synchronously the instant we want to stop — the generator checks
    # this at every segment/character boundary, so it stops queuing further
    # audio immediately rather than waiting on task-cancellation timing
    # (which can be deferred until an in-flight synth call finishes).
    tts_stop_event: asyncio.Event | None = None

    def speak(text: str, trace=None):
        nonlocal tts_task, tts_stop_event
        tts_stop_event = asyncio.Event()
        tts_task = asyncio.create_task(
            _stream_tts_reply(ws, text, tts_fn, _tts_times, trace=trace, stop_event=tts_stop_event))
        return tts_task

    # Speak greeting with claim context right away
    greeting = build_greeting(account)
    conversation.append({"role": "assistant", "content": greeting})
    await ws.send_json({"type": "llm_text", "text": greeting})
    tts_task = speak(greeting)

    vad = VAD(use_silero=use_silero)
    barge_in = False
    full_audio = bytearray()
    call_start = time.time()
    review_saved = False

    async def cancel_tts():
        nonlocal tts_task, tts_stop_event
        if tts_stop_event is not None:
            tts_stop_event.set()
        if tts_task and not tts_task.done():
            tts_task.cancel()
            try:
                await tts_task
            except asyncio.CancelledError:
                pass
            tts_task = None

    async def _finalize_review():
        nonlocal review_saved
        if review_saved:
            return
        review_saved = True
        await _save_review(session_id, conversation, bytes(full_audio),
                           time.time() - call_start, account_uid, call_start,
                           tts_engine, vad_mode, system_prompt,
                           stt_total_ms, stt_count, llm_total_ms, llm_count,
                           sum(_tts_times), len(_tts_times),
                           peak_prompt_tokens, total_completion_tokens)

    try:
        while True:
            raw = await ws.receive_bytes()
            full_audio.extend(raw)
            audio = np.frombuffer(raw, dtype=np.int16).astype(np.float32) / 32768.0
            now = time.time()

            # NOTE: no raw-energy barge-in. The speaker echo of the bot's own
            # TTS frequently exceeded the energy threshold and cancelled
            # replies before they were heard. Interrupts happen only through
            # the VAD segment flush below (Silero distinguishes real speech
            # from the bot's own voice), which still cancels the TTS so the
            # caller can interrupt after finishing their utterance.
            segment = vad.add(audio, now)
            if segment is None:
                continue

            await cancel_tts()
            barge_in = False

            _trace = opik_start_trace("browser.turn", session_id,
                                      {"audio_sec": round(len(segment) / 16000, 2)},
                                      {"type": "browser_call", "session_id": session_id,
                                       "account_uid": account_uid, "state": "browser"})

            _stt_t0 = time.time()
            async with state["stt_lock"]:
                loop = asyncio.get_running_loop()
                text = await loop.run_in_executor(
                    None, lambda: transcribe_live(state["stt_model"], segment))
                stt_latency_ms = (time.time() - _stt_t0) * 1000
                stt_total_ms += stt_latency_ms
                stt_count += 1
            opik_span(_trace, "browser.stt", "general",
                      {"audio_sec": round(len(segment) / 16000, 2)}, {"text": text},
                      _stt_t0, time.time())
            if len(text) < 3:
                opik_end_trace(_trace, output={"text": text, "status": "empty"})
                continue

            if _is_end_of_call(text, conversation):
                await cancel_tts()
                opik_end_trace(_trace, output={"status": "end_of_call"})
                await ws.send_json({"type": "llm_text", "text": "Call ended. Have a great day!"})
                await _finalize_review()
                await asyncio.sleep(1)
                await ws.close()
                return

            conversation.append({"role": "user", "content": f"[INSURANCE REP] {text}"})
            await ws.send_json({"type": "transcript", "text": text})
            conv = conversation if len(conversation) <= 15 else [conversation[0]] + conversation[-14:]
            _llm_t0 = time.time()
            bot_text = ""
            usage = None
            try:
                stream = await state["llm_client"].chat.completions.create(
                    model=state["llm_model"], messages=conv, max_tokens=150, temperature=0,
                    stream=True, stream_options={"include_usage": True})
                async for chunk in stream:
                    if getattr(chunk, "usage", None):
                        usage = chunk.usage
                        continue
                    bot_text += chunk.choices[0].delta.content or ""
            except Exception as _e:
                opik_end_trace(_trace, error=str(_e))
                raise
            llm_total_ms += (time.time() - _llm_t0) * 1000
            llm_count += 1
            if usage is not None:
                peak_prompt_tokens = max(peak_prompt_tokens, int(getattr(usage, "prompt_tokens", 0) or 0))
                total_completion_tokens += int(getattr(usage, "completion_tokens", 0) or 0)

            # Speak the whole response as ONE TTS input (no per-sentence splits,
            # no truncation): even if the model asks two questions, they are
            # spoken together in a single stream.
            cr_idx = bot_text.upper().find("CALL_RESULT")
            cut = bot_text[:cr_idx] if cr_idx != -1 else bot_text
            spoken = strip_markers(cut).rstrip().rstrip("[").strip()
            reply = bot_text.strip()
            result = parse_call_result(reply)
            has_result_marker = bool(result or cr_idx != -1)

            # SAFETY NET: if the response asks a question AND emits [CALL_RESULT],
            # the agent must not finalize before the representative answers —
            # speak the question(s) and wait instead of ending the call.
            if has_result_marker and "?" in spoken:
                log.info("[%s] CALL_RESULT emitted with a pending question — waiting for answer", session_id)
                result = None
                has_result_marker = False

            conversation.append({"role": "assistant", "content": reply})
            opik_span(_trace, "browser.llm", "llm", {"messages": conv[-2:]}, reply,
                      _llm_t0, time.time(), model=state["llm_model"],
                      metadata={"usage": {"prompt_tokens": getattr(usage, "prompt_tokens", None),
                                          "completion_tokens": getattr(usage, "completion_tokens", None),
                                          "total_tokens": getattr(usage, "total_tokens", None)}} if usage else None)

            if result:
                await state["redis"].hset(f"call:{session_id}", mapping={
                    "claim_id": result.get("claim_id") or "browser",
                    "payer": result.get("payer") or "browser",
                    "status": result.get("status") or "completed",
                    "next_action": result.get("next_action") or "",
                    "denial_code": result.get("denial_code") or "",
                    "paid_amount": str(result.get("paid_amount") or ""),
                    "call_summary": result.get("call_summary") or "",
                    "call_result": reply,
                })

            # Speak the whole turn as ONE TTS input.
            if spoken:
                await ws.send_json({"type": "llm_text", "text": spoken})
                tts_task = speak(spoken, trace=_trace)
                if not has_result_marker:
                    # Normal turn: end the trace once this TTS completes.
                    # (the farewell path below ends the trace itself)
                    asyncio.create_task(_end_browser_trace_after_tts(tts_task, _trace, {"reply": spoken}))

            if has_result_marker:
                await cancel_tts()
                if tts_task:
                    try:
                        await tts_task
                    except Exception:
                        pass
                farewell = "Okay, thank you. Goodbye!"
                await ws.send_json({"type": "llm_text", "text": farewell})
                tts_task = speak(farewell, trace=_trace)
                if tts_task:
                    try:
                        await asyncio.wait_for(tts_task, timeout=10)
                    except asyncio.CancelledError:
                        pass
                    except asyncio.TimeoutError:
                        log.warning("[%s] Farewell TTS timed out — ending anyway", session_id)
                    except Exception:
                        pass
                opik_end_trace(_trace, output={"status": "completed", "call_result": result})
                await _finalize_review()
                await asyncio.sleep(1)
                try:
                    await ws.close()
                except Exception:
                    pass
                return

    except Exception:
        pass
    finally:
        await _finalize_review()
        if tts_stop_event is not None:
            tts_stop_event.set()
        if tts_task and not tts_task.done():
            tts_task.cancel()
            try:
                await tts_task
            except asyncio.CancelledError:
                pass


async def _end_browser_trace_after_tts(tts_task, trace, output):
    """End the turn trace after the background TTS (and its span) completes."""
    if trace is None:
        return
    try:
        await tts_task
    except Exception:
        pass
    opik_end_trace(trace, output=output)


async def _stream_tts_reply(ws: WebSocket, text: str, tts_fn=tts_stream, tts_times: list | None = None,
                            trace=None, stop_event: "asyncio.Event | None" = None):
    t0 = time.time()
    first = True
    total_bytes = 0
    try:
        async for pcm_bytes in tts_fn(text, stop_event=stop_event):
            if first and tts_times is not None:
                tts_times.append((time.time() - t0) * 1000)
                first = False
            total_bytes += len(pcm_bytes)
            await ws.send_bytes(b"\x01" + pcm_bytes)
            if first is False and trace is not None:
                opik_span(trace, "browser.tts", "general", {"text": text[:120]},
                          {"audio_bytes": total_bytes, "latency_ms": round((time.time() - t0) * 1000, 1)},
                          t0, time.time())
                trace = None
        if trace is not None:
            opik_span(trace, "browser.tts", "general", {"text": text[:120]},
                      {"audio_bytes": total_bytes}, t0, time.time())
    except asyncio.CancelledError:
        pass
    except Exception as e:
        log.warning("[tts] reply stream error: %s", e)


DASHBOARD_HTML = """<!DOCTYPE html>
<html><head><title>AR Voice Agent — Dashboard</title>
<script src="https://cdn.tailwindcss.com"></script>
<style>@keyframes pulse{0%,100%{opacity:1}50%{opacity:.5}}.live{animation:pulse 2s infinite}</style></head>
<body class="bg-gray-900 text-gray-100 p-6">
<div class="max-w-7xl mx-auto">
  <div class="flex justify-between items-center mb-6">
    <h1 class="text-3xl font-bold">Healthcare AR Voice Agent
      <span class="text-green-400 text-sm live">● ON-PREM</span></h1>
    <div class="flex gap-2">
      <a href="/api/export-excel" class="px-3 py-2 bg-green-700 rounded text-sm font-bold">Export Excel</a>
      <a href="/export.csv" class="px-3 py-2 bg-gray-700 rounded text-sm font-bold">Export CSV</a>
    </div>
  </div>
  <div class="flex gap-2 mb-4">
    <button id="tab-main" onclick="switchTab('main')"
      class="px-4 py-2 rounded text-sm font-bold bg-blue-600">📊 Dashboard</button>
    <button id="tab-review" onclick="switchTab('review')"
      class="px-4 py-2 rounded text-sm font-bold bg-gray-800">🔍 Call Review</button>
  </div>
  <div id="view-main">
  <div class="grid grid-cols-1 lg:grid-cols-3 gap-6">
    <div class="space-y-6">
      <div class="bg-gray-800 p-6 rounded-lg">
        <h2 class="text-lg font-bold mb-4">Outbound Call Control
          <span id="badge" class="text-xs text-blue-400 font-normal block">No account loaded</span></h2>
        <form id="call-form" class="space-y-4">
          <input type="hidden" id="account_uid">
          <input id="phone" placeholder="+15551234567" required
            class="w-full bg-gray-900 border border-gray-700 rounded p-2 text-sm">
          <input id="payer" placeholder="Aetna"
            class="w-full bg-gray-900 border border-gray-700 rounded p-2 text-sm">
          <input id="claim_id" placeholder="CLM-90210"
            class="w-full bg-gray-900 border border-gray-700 rounded p-2 text-sm">
          <button class="w-full bg-blue-600 hover:bg-blue-500 font-bold p-2 rounded text-sm">
            Place Outbound Call</button>
        </form>
        <div id="msg" class="mt-3 text-xs hidden"></div>
      </div>

      <!-- VAD Engine Selector -->
      <div class="bg-gray-800 p-4 rounded-lg">
        <label class="text-xs text-gray-400 block mb-1">VAD Engine</label>
        <select id="vad-engine" class="w-full bg-gray-900 border border-gray-700 rounded p-2 text-sm">
          <option value="silero">Silero (neural)</option>
          <option value="rms">Plain RMS (energy)</option>
        </select>
      </div>

      <!-- TTS Engine Selector -->
      <div class="bg-gray-800 p-4 rounded-lg">
        <label class="text-xs text-gray-400 block mb-1">TTS Engine</label>
        <select id="tts-engine" class="w-full bg-gray-900 border border-gray-700 rounded p-2 text-sm">
          <option value="piper">Piper (fast, robotic)</option>
          <option value="kokoro">Kokoro (natural)</option>
        </select>
      </div>

      <!-- LLM Model Selector -->
      <div class="bg-gray-800 p-4 rounded-lg">
        <label class="text-xs text-gray-400 block mb-1">LLM Model</label>
        <div class="flex gap-2">
          <select id="llm-model" class="flex-1 bg-gray-900 border border-gray-700 rounded p-2 text-sm"></select>
          <button id="switch-llm-btn" onclick="switchLLM()"
            class="px-3 py-1 bg-yellow-700 hover:bg-yellow-600 rounded text-xs font-bold disabled:opacity-50"
            disabled>Switch</button>
        </div>
        <div id="llm-status" class="mt-1 text-xs text-gray-500"></div>
      </div>

      <!-- Browser Call Button -->
      <button id="browser-call-btn" onclick="toggleBrowserCall()"
        class="w-full bg-green-700 hover:bg-green-600 font-bold p-3 rounded text-sm transition duration-200">
        🎤 Call from Browser
      </button>

      <!-- Browser Call Panel (hidden by default) -->
      <div id="browser-call-panel" class="bg-gray-800 rounded-lg p-4 hidden">
        <div class="flex justify-between items-center mb-3">
          <h2 class="text-lg font-bold">🟢 Live Browser Call</h2>
          <button onclick="endBrowserCall()"
            class="px-3 py-1 bg-red-700 hover:bg-red-600 rounded text-xs font-bold">End Call</button>
        </div>
        <div id="browser-log" class="h-48 overflow-y-auto bg-gray-900 rounded p-3 text-xs font-mono space-y-1"></div>
      </div>

      <div class="grid grid-cols-2 gap-4">
        <div class="bg-gray-800 p-4 rounded-lg"><div class="text-sm text-gray-400">Calls</div>
          <div id="calls-today" class="text-2xl font-bold">0</div></div>
        <div class="bg-gray-800 p-4 rounded-lg"><div class="text-sm text-gray-400">Success</div>
          <div id="success-rate" class="text-2xl font-bold">0%</div></div>
        <div class="bg-gray-800 p-4 rounded-lg"><div class="text-sm text-gray-400">Avg Duration</div>
          <div id="avg-dur" class="text-2xl font-bold">0m</div></div>
        <div class="bg-gray-800 p-4 rounded-lg"><div class="text-sm text-gray-400">Cost/min</div>
          <div class="text-2xl font-bold">$0.00</div></div>
      </div>
      <div class="bg-gray-800 rounded-lg p-4">
        <h2 class="text-lg font-bold mb-3">Live Call Feed</h2>
        <div id="call-rows" class="space-y-3 max-h-[300px] overflow-y-auto"></div>
      </div>
    </div>
    <div class="lg:col-span-2 space-y-6">
      <div class="bg-gray-800 p-6 rounded-lg">
        <h2 class="text-lg font-bold mb-2">Excel Calling Context List</h2>
        <form id="upload-form" class="flex gap-4 items-center mt-4">
          <input type="file" id="excel-file" accept=".xlsx" required class="text-sm">
          <button class="px-4 py-2 bg-blue-600 rounded text-sm font-bold">Upload</button>
        </form>
        <div id="upload-msg" class="mt-3 text-xs hidden"></div>
      </div>
      <div class="bg-gray-800 rounded-lg overflow-hidden">
        <div class="p-4 border-b border-gray-700 flex justify-between">
          <h2 class="text-lg font-bold">Calling Checklist</h2>
          <span id="acct-count" class="text-xs text-gray-400">0 Accounts</span></div>
        <div class="overflow-x-auto max-h-[500px] overflow-y-auto">
          <table class="w-full text-sm">
            <thead class="bg-gray-700 sticky top-0"><tr>
              <th class="text-left p-3">Patient</th><th class="text-left p-3">Payer</th>
              <th class="text-left p-3">DOS</th><th class="text-left p-3">Billed</th>
              <th class="text-left p-3">Objective</th><th class="text-left p-3">Outcome</th>
              <th class="text-left p-3">Status</th><th class="text-left p-3"></th></tr></thead>
            <tbody id="acct-rows" class="divide-y divide-gray-700">
              <tr><td colspan="8" class="p-6 text-center text-gray-500">Upload an Excel file.</td></tr>
            </tbody></table></div></div>
    </div>
  </div>
  </div>

  <!-- Call Review View (separate tab) -->
  <div id="view-review" class="hidden">
    <div class="bg-gray-800 rounded-lg p-4">
      <div class="flex justify-between items-center mb-3">
        <h2 class="text-lg font-bold">Call Review
          <span class="text-xs text-gray-400 font-normal">(browser calls only)</span></h2>
        <button onclick="switchTab('main')" class="px-3 py-1 bg-gray-700 rounded text-xs">← Back to Dashboard</button>
      </div>
      <div id="review-list" class="space-y-2 max-h-[400px] overflow-y-auto"></div>
      <div id="review-detail" class="hidden mt-4 border-t border-gray-700 pt-4">
        <div class="flex justify-between items-center mb-3">
          <h3 class="text-md font-bold">Transcript Comparison</h3>
          <button onclick="closeReview()" class="px-2 py-1 bg-gray-700 rounded text-xs">✕ Close</button>
        </div>
        <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
          <div class="bg-gray-900 rounded p-3">
            <div class="text-xs text-blue-300 font-bold mb-2">🔴 Real-Time STT (with VAD)</div>
            <div id="review-real-time" class="text-xs leading-relaxed whitespace-pre-wrap"></div>
          </div>
          <div class="bg-gray-900 rounded p-3">
            <div class="text-xs text-green-300 font-bold mb-2">🟢 Full Recording STT (no VAD)</div>
            <div id="review-full-audio" class="text-xs leading-relaxed whitespace-pre-wrap"></div>
          </div>
        </div>
        <div class="mt-3 bg-gray-900 rounded p-3">
          <div class="text-xs text-orange-300 font-bold mb-2">🤖 AI Responses</div>
          <div id="review-ai-responses" class="text-xs leading-relaxed whitespace-pre-wrap"></div>
        </div>
        <div class="mt-2 text-[10px] text-gray-500" id="review-meta"></div>
      </div>
    </div>
  </div>
</div>
<script>
const $ = id => document.getElementById(id);
let stats = {total: 0, success: 0};

async function fetchCalls() {
  const calls = await (await fetch('/api/calls')).json();
  $('call-rows').innerHTML = '';
  stats = {total: 0, success: 0};
  let durSum = 0, durN = 0;
  calls.forEach(c => {
    stats.total++;
    if (c.status === 'completed') stats.success++;
    if (c.duration_ms) { durSum += +c.duration_ms; durN++; }
    const color = c.status==='completed'?'bg-green-900 text-green-300'
      : c.status==='failed'?'bg-red-900 text-red-300':'bg-yellow-900 text-yellow-300';
    $('call-rows').insertAdjacentHTML('afterbegin', `
      <div class="p-3 bg-gray-900 rounded border border-gray-700">
        <div class="flex justify-between"><b class="text-xs">${c.payer||'Unknown'}</b>
        <span class="px-1.5 py-0.5 rounded text-[10px] ${color}">${c.status}</span></div>
        <div class="text-[11px] text-gray-400 mt-1">Claim: ${c.claim_id||'-'}
          ${c.amount?'<br>Billed: $'+c.amount:''}
          ${c.next_action?'<br>Action: '+c.next_action:''}
          ${c.last_error?'<br><span class="text-red-400">'+c.last_error+'</span>':''}
        </div></div>`);
  });
  $('calls-today').textContent = stats.total;
  $('success-rate').textContent = stats.total ? Math.round(stats.success/stats.total*100)+'%' : '0%';
  $('avg-dur').textContent = durN ? Math.round(durSum/durN/60000*10)/10+'m' : '0m';
}

async function fetchAccounts() {
  const accts = await (await fetch('/api/accounts')).json();
  $('acct-count').textContent = accts.length + ' Accounts';
  if (!accts.length) return;
  $('acct-rows').innerHTML = '';
  accts.forEach(a => {
    const st = a['Call Status']||'Pending';
    const color = st==='Calls Done'?'bg-green-950 text-green-300'
      : (st==='Failed'||st==='Disconnected')?'bg-red-950 text-red-300':'bg-gray-700';
    $('acct-rows').insertAdjacentHTML('beforeend', `<tr class="hover:bg-gray-700">
      <td class="p-3 font-semibold">${a['Patient Name']||'-'}</td>
      <td class="p-3">${a['Responsible Payer']||''}</td>
      <td class="p-3 text-gray-400">${(a['DOS']||'').slice(0,10)}</td>
      <td class="p-3">${a['Billed Amount']?'$'+a['Billed Amount']:'-'}</td>
      <td class="p-3 text-xs text-gray-400 max-w-xs truncate">${a['AR Final Comments']||'-'}</td>
      <td class="p-3 text-xs max-w-xs truncate">${a['Call Comments']||'-'}</td>
      <td class="p-3"><span class="px-2 py-0.5 rounded text-[10px] font-bold ${color}">${st}</span></td>
      <td class="p-3"><button onclick="pick('${a.UID}','${(a['Patient Name']||'').replace(/'/g,"")}','${a['Responsible Payer']||''}','${a['Account Number']||''}')"
        class="px-2 py-1 bg-blue-600 rounded text-xs font-bold">Load</button></td></tr>`);
  });
}

function pick(uid, name, payer, claim) {
  $('account_uid').value = uid; $('payer').value = payer; $('claim_id').value = claim;
  $('badge').textContent = 'Loaded: ' + name;
}

$('upload-form').onsubmit = async e => {
  e.preventDefault();
  const fd = new FormData();
  fd.append('file', $('excel-file').files[0]);
  const res = await fetch('/api/upload-excel', {method: 'POST', body: fd});
  const d = await res.json();
  $('upload-msg').classList.remove('hidden');
  $('upload-msg').textContent = res.ok ? `Loaded ${d.count} accounts` : 'Error: ' + d.error;
  $('upload-msg').className = 'mt-3 text-xs ' + (res.ok ? 'text-green-400' : 'text-red-400');
  fetchAccounts();
};

$('call-form').onsubmit = async e => {
  e.preventDefault();
  const res = await fetch('/make-call', {method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({phone: $('phone').value, payer: $('payer').value,
      claim_id: $('claim_id').value, account_uid: $('account_uid').value})});
  const d = await res.json();
  $('msg').classList.remove('hidden');
  $('msg').textContent = res.ok ? 'Call triggered: ' + d.callSid : 'Error: ' + d.error;
  $('msg').className = 'mt-3 text-xs ' + (res.ok ? 'text-green-400' : 'text-red-400');
};

fetchAccounts(); fetchCalls();
setInterval(fetchCalls, 3000);
setInterval(fetchAccounts, 5000);

// ── Browser Call ──────────────────────────────────────────────────────
let bcWS = null, bcMic = null, bcCtx = null, bcSrc = null, bcProc = null;
let bcPlayCtx = null, bcAudioQ = [], bcPlaying = false, bcCurSrc = null;
let bcLogEl = null;

function bcLog(msg, cls) {
  const d = document.createElement('div');
  d.textContent = msg;
  if (cls) d.className = cls;
  if (bcLogEl) { bcLogEl.appendChild(d); bcLogEl.scrollTop = bcLogEl.scrollHeight; }
}

function toggleBrowserCall() {
  if (bcWS && bcWS.readyState === WebSocket.OPEN) { endBrowserCall(); return; }
  startBrowserCall();
}

function endBrowserCall() {
  if (bcPlayCtx) { bcPlayCtx.close(); bcPlayCtx = null; }
  if (bcProc) { bcProc.disconnect(); bcProc = null; }
  if (bcSrc) { bcSrc.disconnect(); bcSrc = null; }
  if (bcMic) { bcMic.getTracks().forEach(t => t.stop()); bcMic = null; }
  if (bcCtx) { bcCtx.close(); bcCtx = null; }
  if (bcWS) { bcWS.close(); bcWS = null; }
  bcAudioQ = []; bcPlaying = false; bcCurSrc = null;
  $('browser-call-panel').classList.add('hidden');
  $('browser-call-btn').textContent = '🎤 Call from Browser';
  $('browser-call-btn').className = 'w-full bg-green-700 hover:bg-green-600 font-bold p-3 rounded text-sm';
}

async function startBrowserCall() {
  const sid = 'browser_' + Math.random().toString(36).slice(2);
  const uid = ($('account_uid') && $('account_uid').value) || '';
  const tts = ($('tts-engine') && $('tts-engine').value) || 'piper';
  const vad = ($('vad-engine') && $('vad-engine').value) || 'silero';
  const params = new URLSearchParams();
  if (uid) params.set('account_uid', uid);
  params.set('tts', tts);
  params.set('vad', vad);
  bcPlayCtx = new AudioContext(); // create inside user gesture — ensures running state
  bcWS = new WebSocket('wss://' + location.host + '/ws/' + sid + '?' + params.toString());
  bcWS.binaryType = 'arraybuffer';

  $('browser-call-panel').classList.remove('hidden');
  bcLogEl = $('browser-log');
  bcLogEl.innerHTML = '<div class="text-green-400">Connecting...</div>';
  $('browser-call-btn').textContent = '🔴 End Browser Call';
  $('browser-call-btn').className = 'w-full bg-red-700 hover:bg-red-600 font-bold p-3 rounded text-sm';

  bcWS.onopen = async () => {
    bcLog('Connected — starting mic...', 'text-green-400');
    try {
      bcCtx = new AudioContext({ sampleRate: 16000 });
      bcMic = await navigator.mediaDevices.getUserMedia({ audio: true });
      bcSrc = bcCtx.createMediaStreamSource(bcMic);
      bcProc = bcCtx.createScriptProcessor(4096, 1, 1);
      bcProc.onaudioprocess = e => {
        if (!bcWS || bcWS.readyState !== 1) return;
        const inp = e.inputBuffer.getChannelData(0);
        const b = new Int16Array(inp.length);
        for (let i = 0; i < inp.length; i++) b[i] = Math.max(-32768, Math.min(32767, inp[i] * 32768));
        bcWS.send(b.buffer);
      };
      bcSrc.connect(bcProc);
      bcProc.connect(bcCtx.destination);
      bcLog('✅ Mic active — speak now', 'text-green-400');
    } catch (e) {
      bcLog('❌ Mic error: ' + e.message, 'text-red-400');
    }
  };

  bcWS.onmessage = e => {
    if (typeof e.data === 'string') {
      const m = JSON.parse(e.data);
      if (m.type === 'config') { bcPlayCtx = new AudioContext(); return; }
      bcLog(m.text, m.type === 'transcript' ? 'text-blue-300' : 'text-orange-300');
    } else {
      const v = new Uint8Array(e.data);
      if (v[0] === 1) { bcAudioQ.push(v.slice(1)); if (!bcPlaying) bcPlayNext(); }
      else if (v[0] === 2) { bcAudioQ = []; if (bcPlaying && bcCurSrc) { bcPlaying = false; try { bcCurSrc.stop(); } catch (e) {} } }
    }
  };

  bcWS.onclose = () => {
    bcLogEl.innerHTML += '<div class="text-gray-500">Disconnected.</div>';
    endBrowserCall();
  };
}

function bcPlayNext() {
  if (!bcAudioQ.length || !bcPlayCtx) { bcPlaying = false; return; }
  bcPlaying = true;
  const total = bcAudioQ.reduce((s, c) => s + c.length, 0);
  const pcm = new Int16Array(total / 2); let off = 0;
  while (bcAudioQ.length) { const c = bcAudioQ.shift(); pcm.set(new Int16Array(c.buffer, c.byteOffset, c.length / 2), off); off += c.length / 2; }
  const buf = bcPlayCtx.createBuffer(1, pcm.length, 22050);
  const ch = buf.getChannelData(0);
  for (let i = 0; i < pcm.length; i++) ch[i] = pcm[i] / 32768;
  const s = bcPlayCtx.createBufferSource();
  bcCurSrc = s;
  s.buffer = buf; s.connect(bcPlayCtx.destination);
  s.onended = () => { bcPlaying = false; bcCurSrc = null; if (bcAudioQ.length) bcPlayNext(); };
  s.start();
}

// ── LLM Model Switcher ──────────────────────────────────────────────
async function loadLLM() {
  const sel = $('llm-model');
  const st = $('llm-status');
  const btn = $('switch-llm-btn');
  try {
    const r = await (await fetch('/api/current-llm')).json();
    sel.innerHTML = Object.keys(r.options).map(l =>
      `<option value="${r.options[l]}" ${r.options[l]===r.model?'selected':''}>${l}</option>`
    ).join('');
    btn.disabled = false;
    st.textContent = r.switching ? '🔄 Switching...' : r.model;
    if (r.switching) setTimeout(loadLLM, 2000);
  } catch (e) { st.textContent = 'Failed to load'; }
}
async function switchLLM() {
  const sel = $('llm-model');
  const st = $('llm-status');
  const btn = $('switch-llm-btn');
  btn.disabled = true;
  st.textContent = '🔄 Switching model, vLLM restarting (~1 min)...';
  try {
    const r = await (await fetch('/api/switch-llm', {method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({model: sel.value})
    })).json();
    if (r.error) { st.textContent = 'Error: ' + r.error; btn.disabled = false; return; }
    setTimeout(loadLLM, 3000);
  } catch (e) { st.textContent = 'Error: ' + e.message; btn.disabled = false; }
}
loadLLM();

// ── Review Tab ──────────────────────────────────────────────────────
function switchTab(name) {
  const main = name === 'main';
  $('view-main').classList.toggle('hidden', !main);
  $('view-review').classList.toggle('hidden', main);
  $('tab-main').className = 'px-4 py-2 rounded text-sm font-bold ' + (main ? 'bg-blue-600' : 'bg-gray-800');
  $('tab-review').className = 'px-4 py-2 rounded text-sm font-bold ' + (main ? 'bg-gray-800' : 'bg-blue-600');
  if (!main) fetchReviews();
}

function highlightDiff(realText, fullText) {
  const realWords = (realText || '').split(/\\s+/).filter(Boolean);
  const fullWords = (fullText || '').split(/\\s+/).filter(Boolean);
  const fullSet = new Set(fullWords.map(w => w.toLowerCase()));
  const realSet = new Set(realWords.map(w => w.toLowerCase()));
  const missed = realWords.filter(w => !fullSet.has(w.toLowerCase()));
  const extra = fullWords.filter(w => !realSet.has(w.toLowerCase()));

  const highlight = (words, highlights, cls) =>
    words.map(w => highlights.some(h => h.toLowerCase() === w.toLowerCase())
      ? `<span class="${cls}">${w}</span>` : w).join(' ');

  const hReal = highlight(realWords, missed, 'text-red-400 font-bold');
  const hFull = highlight(fullWords, extra, 'text-green-400 font-bold');
  return { hReal, hFull, missed: missed.length, extra: extra.length };
}

async function fetchReviews() {
  try {
    const calls = await (await fetch('/api/calls')).json();
    const callable = calls.filter(c => c.callSid && (c.callSid.startsWith('browser_') || c.callSid.startsWith('CA')));
    const list = $('review-list');
    if (!callable.length) {
      list.innerHTML = '<div class="text-gray-500 text-xs">No calls yet</div>';
      return;
    }
    list.innerHTML = callable.slice(0, 10).map(c =>
      `<div class="flex justify-between items-center bg-gray-900 rounded p-2 text-xs">
        <span class="text-gray-400">${c.callSid.slice(0, 24)}…</span>
        <span class="${c.status==='completed'?'text-green-400':'text-yellow-400'}">${c.status}</span>
        <button onclick="openReview('${c.callSid}')"
          class="px-2 py-0.5 bg-blue-700 hover:bg-blue-600 rounded text-[10px]">Review</button>
      </div>`
    ).join('');
  } catch (e) {}
}

async function openReview(sid) {
  const detail = $('review-detail');
  const rTime = $('review-real-time');
  const rFull = $('review-full-audio');
  const rAI = $('review-ai-responses');
  const rMeta = $('review-meta');
  detail.classList.remove('hidden');
  rTime.textContent = 'Loading...';
  rFull.textContent = '';
  rAI.textContent = '';
  rMeta.textContent = '';
  try {
    const d = await (await fetch(`/api/calls/${sid}/review`)).json();
    if (d.error) {
      if (sid.startsWith('browser_')) {
        rTime.textContent = '⏳ Offline transcription in progress — refresh in a few seconds.';
        setTimeout(() => openReview(sid), 5000);
      } else {
        rTime.textContent = 'No transcript available for this call.';
      }
      return;
    }

    const userText = (d.real_time || []).join('\\n');
    const fullText = d.full_audio || '';

    const diff = highlightDiff(userText, fullText);
    rTime.innerHTML = diff.hReal || '<span class="text-gray-500">(no speech detected)</span>';
    rFull.innerHTML = diff.hFull || '<span class="text-gray-500">(no speech detected)</span>';
    rAI.textContent = (d.ai_responses || []).join('\\n') || '(none)';

    const dur = d.duration_sec ? Math.round(d.duration_sec / 60 * 10) / 10 + 'm' : '?';
    const size = d.audio_size_bytes ? (d.audio_size_bytes / 1024 / 1024).toFixed(1) + 'MB' : '?';
    const words = fullText ? fullText.split(/\\s+/).length : 0;
    rMeta.textContent = `Duration: ${dur} · Audio: ${size} · Words: ${words} · ` +
      `Missed: ${diff.missed} · Extra: ${diff.extra}`;
  } catch (e) { rTime.textContent = 'Error loading review'; }
}

function closeReview() {
  $('review-detail').classList.add('hidden');
}

fetchReviews();
setInterval(fetchReviews, 5000);
</script></body></html>"""


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
