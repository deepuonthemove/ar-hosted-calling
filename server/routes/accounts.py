"""Excel upload, projects, accounts, notes, and LLM-context overrides."""
import io
import json
import time
import uuid

from fastapi import APIRouter, File, Request, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

from prompts import build_call_prompt

from ..config import log
from ..state import state
from .stats import _call_record_keys

router = APIRouter()

# ── Excel header mapping ─────────────────────────────────────────────────
HEADER_ALIASES = {
    # objective
    "final comments": "AR Final Comments",
    "ar final comments": "AR Final Comments",
    # patient / demographics
    "patient": "Patient Name",
    "patient name": "Patient Name",
    "patient's name": "Patient Name",
    "member id": "Member ID",
    "member": "Member ID",
    "dob": "DOB",
    "date of birth": "DOB",
    "dos": "DOS",
    "date of service": "DOS",
    "cpt": "CPT",
    "cpt code": "CPT",
    "billed amount": "Billed Amount",
    "billed": "Billed Amount",
    "amount billed": "Billed Amount",
    # payer / billing
    "responsible party": "Responsible Payer",
    "responsible payer": "Responsible Payer",
    "payer": "Responsible Payer",
    "account number": "Account Number",
    "account #": "Account Number",
    "claim id": "Claim ID",
    "claim number": "Claim ID",
    # group / provider
    "group": "Group",
    "group name": "Group Name",
    "group npi": "Group NPI",
    "tax id": "Tax ID",
    "billing/provider address": "Billing/Provider Address",
    "billing address": "Billing/Provider Address",
    "provider address": "Billing/Provider Address",
    "pay-to-address": "Pay-to-address",
    "pay to address": "Pay-to-address",
    "individual npi": "Individual NPI",
    # workflow / outcome
    "call status": "Call Status",
    "status": "Call Status",
    "denial code": "Denial Code",
    "denial": "Denial Code",
    "notes": "Notes",
    "uid": "UID",
}


def _normalize_header(h: str) -> str:
    """Map a source Excel header to the canonical field name (case-insensitive)."""
    key = h.strip().lower()
    return HEADER_ALIASES.get(key, h.strip())


# ════════════════════════════════════════════════════════════════════════
# EXCEL  (ported from src/index.ts upload/accounts/export)
# ════════════════════════════════════════════════════════════════════════

async def _resolve_account(project_id: str, row_num: int) -> tuple[str | None, dict | None]:
    """Resolve (project_id, row_num) → (account_uid, account_record). Row 1 = first data row."""
    raw = await state["redis"].get(f"project:{project_id}:rows")
    if not raw:
        return None, None
    uids = json.loads(raw)
    if not (1 <= row_num <= len(uids)):
        return None, None
    uid = uids[row_num - 1]
    acct = await state["redis"].hgetall(f"account:{uid}") or None
    return uid, acct


async def _effective_prompt(account: dict | None) -> str:
    """System prompt for an account: custom override if set, else the built default.
    Notes are appended in either case."""
    if not account:
        return build_call_prompt("GREETING", None, None, None)
    uid = account.get("UID", "")
    custom = None
    if uid:
        custom = await state["redis"].get(f"account:{uid}:llm_context")
    notes = account.get("Notes", "")
    if custom:
        prompt = custom
        if notes:
            prompt += f"\n\n[PRIOR CALL NOTES]\n{notes}"
        return prompt
    return build_call_prompt("GREETING", None, None, account)


@router.get("/api/accounts/{account_uid}/llm-context")
async def get_llm_context(account_uid: str):
    account = await state["redis"].hgetall(f"account:{account_uid}") or None
    custom = await state["redis"].get(f"account:{account_uid}:llm_context") or ""
    original = build_call_prompt("GREETING", None, None, account)
    effective = await _effective_prompt(account)
    return {"account_uid": account_uid, "custom": custom, "original": original, "effective": effective}


@router.post("/api/accounts/{account_uid}/llm-context")
async def update_llm_context(account_uid: str, request: Request):
    data = await request.json()
    key = f"account:{account_uid}:llm_context"
    if data.get("reset"):
        await state["redis"].delete(key)
    else:
        ctx = (data.get("context") or "").strip()
        if not ctx:
            return JSONResponse({"error": "context is required"}, 400)
        await state["redis"].set(key, ctx)
    account = await state["redis"].hgetall(f"account:{account_uid}") or None
    return {"ok": True, "custom": (await state["redis"].get(key)) or "",
            "effective": await _effective_prompt(account)}


@router.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return JSONResponse({"error": "Sheet is empty"}, 400)

        headers = [_normalize_header(str(h) if h else f"col_{i}") for i, h in enumerate(rows[0])]
        r = state["redis"]

        # Generate a unique project id for this upload
        project_id = f"proj-{uuid.uuid4().hex[:10]}"

        uids = []
        for i, row in enumerate(rows[1:]):
            record = {}
            for j, val in enumerate(row):
                if j < len(headers) and val is not None:
                    record[headers[j]] = val.isoformat() if hasattr(val, "isoformat") else str(val)
            # "Final Comments" → objective (AR Final Comments) is handled by HEADER_ALIASES
            uid = record.get("UID") or f"{project_id}-{i}-{int(time.time())}"
            record["UID"] = uid
            record.setdefault("Call Status", "Pending")
            uids.append(uid)
            await r.hset(f"account:{uid}", mapping=record)

        # Project-scoped row index: index 0 = row 1 = first data row
        await r.set(f"project:{project_id}:rows", json.dumps(uids))
        await r.set(f"project:{project_id}:headers", json.dumps(headers))
        # Keep legacy lists for the UI (last project shown)
        await r.set("accounts-headers", json.dumps(headers))
        await r.set("accounts-list", json.dumps(uids))
        return {"ok": True, "count": len(uids), "project_id": project_id}
    except Exception as e:
        log.error("Excel parse error: %s", e)
        return JSONResponse({"error": f"Parsing error: {e}"}, 500)


@router.get("/api/projects/{project_id}/accounts/{row_num}")
async def get_project_account(project_id: str, row_num: int):
    """Resolve an account by project + row number (row 1 = first data row)."""
    uid, acct = await _resolve_account(project_id, row_num)
    if not acct:
        return JSONResponse({"error": f"No account for project {project_id} row {row_num}"}, 404)
    return {"account_uid": uid, "account": acct}


@router.get("/api/projects")
async def list_projects():
    """List all uploaded projects with row counts."""
    keys = await state["redis"].keys("project:*:rows")
    projects = []
    for key in keys:
        pid = key.replace("project:", "").replace(":rows", "")
        raw = await state["redis"].get(key)
        try:
            count = len(json.loads(raw)) if raw else 0
        except json.JSONDecodeError:
            count = 0
        projects.append({"project_id": pid, "rows": count})
    projects.sort(key=lambda p: p["project_id"], reverse=True)
    return projects


@router.delete("/api/projects/{project_id}")
async def delete_project(project_id: str):
    """Delete a project: its metadata, accounts, and the call records they own."""
    r = state["redis"]
    raw = await r.get(f"project:{project_id}:rows")
    if not raw:
        return JSONResponse({"error": f"Project {project_id} not found"}, 404)
    uids = json.loads(raw)
    uid_set = set(uids)

    await r.delete(f"project:{project_id}:rows", f"project:{project_id}:headers")

    for uid in uids:
        await r.delete(f"account:{uid}")
        await r.delete(f"account:{uid}:llm_context")
        await r.delete(f"account:{uid}:active")
        await r.delete(f"account:{uid}:active:meta")

    # Remove call records (and sub-keys) that belong to this project's accounts
    call_keys = await r.keys("call:*")
    base_keys = _call_record_keys(call_keys)
    deleted_calls = 0
    for key in base_keys:
        data = await r.hgetall(key)
        if data.get("account_uid") in uid_set:
            base = key.replace("call:", "")
            await r.delete(key)
            for suffix in (":audio", ":review", ":transcript", ":meta", ":live"):
                await r.delete(f"call:{base}{suffix}")
            deleted_calls += 1

    return {"ok": True, "deleted_accounts": len(uids), "deleted_calls": deleted_calls}


@router.get("/api/projects/{project_id}/accounts")
async def list_project_accounts(project_id: str):
    raw = await state["redis"].get(f"project:{project_id}:rows")
    if not raw:
        return []
    uids = json.loads(raw)
    accounts = []
    for uid in uids:
        row = await state["redis"].hgetall(f"account:{uid}")
        if row:
            accounts.append(row)
    return accounts


@router.post("/api/notes")
async def add_note(request: Request):
    """Add a note to an account (used as context on the next call).
    Body: {account_uid} or {project_id, row_num} plus {note}."""
    data = await request.json()
    note = (data.get("note") or "").strip()
    if not note:
        return JSONResponse({"error": "note is required"}, 400)

    uid = data.get("account_uid", "")
    if not uid and data.get("project_id") and data.get("row_num"):
        uid, _ = await _resolve_account(data["project_id"], int(data["row_num"]))
    if not uid:
        return JSONResponse({"error": "account_uid or project_id+row_num required"}, 400)

    account = await state["redis"].hgetall(f"account:{uid}") or {}
    prior = account.get("Notes", "")
    now = time.strftime("%Y-%m-%d %H:%M")
    merged = (prior + "\n" if prior else "") + f"[{now}] {note}"
    await state["redis"].hset(f"account:{uid}", mapping={"Notes": merged})
    return {"ok": True, "account_uid": uid, "notes": merged}
